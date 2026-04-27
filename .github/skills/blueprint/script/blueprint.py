#!/usr/bin/env python3
"""
blueprint.py - NVL Blueprint Excel reader and updater.

Fixed Excel path (relative to repo root):
    Shared/BaseInputs/Common/Common_Files/NVL_Blueprint.xlsx

This script is located at:
    .github/skills/blueprint/script/blueprint.py
Repo root is resolved as 4 directories up from this file.

Usage:
  python blueprint.py --read [--sheet <name>]
  python blueprint.py --query <text> [--sheet <name>]
  python blueprint.py --update-cell --sheet <name> --row <n> --col <col> --value <val>
  python blueprint.py --update-match --sheet <name> --match-col <col> --match-val <val> --update-col <col> --update-val <val>

All output is JSON to stdout.
Before any write operation, a .bak copy of the Excel file is created automatically.
"""

import argparse
import copy
import json
import os
import shutil
import sys
from pathlib import Path

from programflows_sync import apply_sync_plan, build_sync_plan, serialize_plan

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string, get_column_letter
except ImportError:
    print(json.dumps({"error": "openpyxl is not installed. Run: pip install openpyxl"}))
    sys.exit(1)


# ---------------------------------------------------------------------------
# Path resolution
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
# Script lives at: <repo_root>/.github/skills/blueprint/script/blueprint.py
# So repo root is 4 levels up: script -> blueprint -> skills -> .github -> repo_root
REPO_ROOT = SCRIPT_DIR.parent.parent.parent.parent
EXCEL_PATH = REPO_ROOT / "Shared" / "BaseInputs" / "Common" / "Common_Files" / "NVL_Blueprint.xlsx"
ALLOWED_SHEETS = ["SimplerBPView", "FullBP"]
DIELET_KEYS = ["CPU", "GCD", "HUB", "PCD"]
# Semantic aliases for structurally meaningful rows.
# FullBP row coverage itself is dynamic and comes from column-A discovery.
SEMANTIC_ROW_LABELS = {
    "tos_exec_flow": "TOS_EXEC_FLOW",
    "tp_exec_flow": "TP_EXEC_FLOW",
    "top_flow": "TOP_FLOW:Flow",
    "top_flow_type": "TOP_FLOW:FlowType",
    "top_port_range": "TOP_FLOW:PortRange",
    "top_dielet_type": "TOP_FLOW:DieletType",
    "all_flow": "FLOWDEF:ALL:Flow",
    "all_dielet_type": "FLOWDEF:ALL:DieletType",
    "all_scope": "FLOWDEF:ALL:Scope",
    "all_port_range": "FLOWDEF:ALL:PortRange",
    "all_flow_type": "FLOWDEF:ALL:FlowType",
    "cpu_flow": "FLOWDEF:CPU:Flow",
    "gcd_flow": "FLOWDEF:GCD:Flow",
    "hub_flow": "FLOWDEF:HUB:Flow",
    "pcd_flow": "FLOWDEF:PCD:Flow",
}
DIELET_FIELD_MAP = {
    "CPU": "cpu_flow",
    "GCD": "gcd_flow",
    "HUB": "hub_flow",
    "PCD": "pcd_flow",
}
DIELET_METADATA_KEYS = ["scope", "flow_type", "port_range"]
for dielet in DIELET_KEYS:
    row_prefix = dielet.lower()
    SEMANTIC_ROW_LABELS[f"{row_prefix}_scope"] = f"FLOWDEF:{dielet}:Scope"
    SEMANTIC_ROW_LABELS[f"{row_prefix}_flow_type"] = f"FLOWDEF:{dielet}:FlowType"
    SEMANTIC_ROW_LABELS[f"{row_prefix}_port_range"] = f"FLOWDEF:{dielet}:PortRange"
SEMANTIC_FIELD_BY_ROW_LABEL = {
    row_label: field_name for field_name, row_label in SEMANTIC_ROW_LABELS.items()
}
TOP_LEVEL_FIELD_KEYS = [
    "tos_exec_flow",
    "tp_exec_flow",
    "top_flow",
    "top_flow_type",
    "top_port_range",
    "top_dielet_type",
]
PLAN_VALUE_STATUSES = {"explicit", "inferred", "copied_from_template"}
SERIAL_FLOW_TYPE = "SubFlow"
PARALLEL_FLOW_TYPE = "SubParFlow"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def resolve_col(col_str):
    """Accept a column letter (A, B, ...) or 1-based integer string. Return 1-based int."""
    try:
        idx = int(col_str)
        if idx < 1:
            raise ValueError("Column index must be >= 1")
        return idx
    except ValueError:
        pass
    try:
        return column_index_from_string(col_str.strip().upper())
    except Exception:
        raise ValueError(
            f"Invalid column: '{col_str}'. Use a letter (A, B, ...) or a 1-based integer."
        )


def get_cell_color(cell):
    """Extract background fill color as an ARGB hex string, or None if no meaningful fill."""
    try:
        fill = cell.fill
        if fill is None:
            return None
        fill_type = fill.fill_type
        if fill_type in (None, "none"):
            return None
        fg = fill.fgColor
        if fg.type == "rgb":
            rgb = fg.rgb.upper()
            # Filter out transparent/default placeholder values
            if rgb in ("00000000", "FFFFFFFF", "FF000000", "00FFFFFF"):
                return None
            return rgb
        if fg.type == "theme":
            tint = getattr(fg, "tint", 0.0)
            return f"theme:{fg.theme},tint:{tint}"
        # indexed colors — skip (palette-dependent, not portable)
        return None
    except Exception:
        return None


def coerce_value(val_str):
    """Try to coerce a string to int, then float, then leave as str."""
    try:
        return int(val_str)
    except (ValueError, TypeError):
        pass
    try:
        return float(val_str)
    except (ValueError, TypeError):
        pass
    return val_str


def normalize_text(value):
    """Normalize a cell value for matching/comparison."""
    if value is None:
        return ""
    return str(value).strip()


def get_excel_path():
    """Return the configured workbook path, allowing an env-var override for safe validation."""
    override = os.environ.get("BLUEPRINT_EXCEL_PATH")
    if override:
        return Path(override).expanduser().resolve()
    return EXCEL_PATH


def load_workbook_readonly(path):
    """Load workbook in data_only mode (for read/query commands)."""
    if not path.exists():
        print(json.dumps({"error": f"Excel file not found: {path}"}))
        sys.exit(1)
    return openpyxl.load_workbook(str(path), data_only=True)


def load_workbook_writable(path):
    """Load workbook preserving formulas (for update commands)."""
    if not path.exists():
        print(json.dumps({"error": f"Excel file not found: {path}"}))
        sys.exit(1)
    return openpyxl.load_workbook(str(path))


def get_sheet(wb, sheet_name):
    """Return the named worksheet, or the active sheet if sheet_name is None."""
    if sheet_name is None:
        return wb[ALLOWED_SHEETS[0]]
    if sheet_name not in ALLOWED_SHEETS:
        print(json.dumps({
            "error": f"Sheet '{sheet_name}' is not supported.",
            "allowed_sheets": ALLOWED_SHEETS
        }))
        sys.exit(1)
    if sheet_name not in wb.sheetnames:
        print(json.dumps({
            "error": f"Sheet '{sheet_name}' not found.",
            "available_sheets": wb.sheetnames
        }))
        sys.exit(1)
    return wb[sheet_name]


def sheet_to_rows(ws):
    """Convert a worksheet to a list of row dicts with cell metadata."""
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        cells = []
        for cell in row:
            col_letter = get_column_letter(cell.column)
            cells.append({
                "col": col_letter,
                "col_index": cell.column,
                "value": cell.value,
                "color": get_cell_color(cell)
            })
        rows.append({"row": row_idx, "cells": cells})
    return rows


def build_row_lookup(ws):
    """Map row label in column A to its row number."""
    lookup = {}
    for row_idx in range(1, ws.max_row + 1):
        label = normalize_text(ws.cell(row=row_idx, column=1).value)
        if label:
            lookup[label] = row_idx
    return lookup


def get_cell_payload(ws, row_idx, col_idx):
    """Return structured cell metadata or None for an empty cell."""
    if row_idx is None:
        return None
    cell = ws.cell(row=row_idx, column=col_idx)
    if cell.value is None:
        return None
    return {
        "value": cell.value,
        "color": get_cell_color(cell),
        "row": row_idx,
        "col": get_column_letter(col_idx),
        "col_index": col_idx,
    }


def list_fullbp_row_labels(ws):
    """Return all non-empty row labels from column A in worksheet order."""
    labels = []
    for row_idx in range(1, ws.max_row + 1):
        label = normalize_text(ws.cell(row=row_idx, column=1).value)
        if label:
            labels.append(label)
    return labels


def build_slot_view(ws, fullbp_index=None, include_fullbp_rows=False):
    """Build a slot-oriented view anchored on FLOWDEF:ALL:Flow for human-readable blueprint ops."""
    row_lookup = build_row_lookup(ws)
    all_flow_row = row_lookup.get(SEMANTIC_ROW_LABELS["all_flow"])
    if all_flow_row is None:
        return []

    fullbp_row_labels = list(row_lookup.keys()) if include_fullbp_rows and ws.title == "FullBP" else []

    slots = []
    for col_idx in range(2, ws.max_column + 1):
        all_flow = get_cell_payload(ws, all_flow_row, col_idx)
        if all_flow is None:
            continue

        slot = {
            "sheet": ws.title,
            "slot_col": get_column_letter(col_idx),
            "slot_col_index": col_idx,
            "tos_exec_flow": get_cell_payload(ws, row_lookup.get(SEMANTIC_ROW_LABELS["tos_exec_flow"]), col_idx),
            "tp_exec_flow": get_cell_payload(ws, row_lookup.get(SEMANTIC_ROW_LABELS["tp_exec_flow"]), col_idx),
            "top_flow": get_cell_payload(ws, row_lookup.get(SEMANTIC_ROW_LABELS["top_flow"]), col_idx),
            "top_flow_type": get_cell_payload(ws, row_lookup.get(SEMANTIC_ROW_LABELS["top_flow_type"]), col_idx),
            "top_port_range": get_cell_payload(ws, row_lookup.get(SEMANTIC_ROW_LABELS["top_port_range"]), col_idx),
            "top_dielet_type": get_cell_payload(ws, row_lookup.get(SEMANTIC_ROW_LABELS["top_dielet_type"]), col_idx),
            "all_flow": all_flow,
            "all_dielet_type": get_cell_payload(ws, row_lookup.get(SEMANTIC_ROW_LABELS["all_dielet_type"]), col_idx),
            "all_scope": get_cell_payload(ws, row_lookup.get(SEMANTIC_ROW_LABELS["all_scope"]), col_idx),
            "all_port_range": get_cell_payload(ws, row_lookup.get(SEMANTIC_ROW_LABELS["all_port_range"]), col_idx),
            "all_flow_type": get_cell_payload(ws, row_lookup.get(SEMANTIC_ROW_LABELS["all_flow_type"]), col_idx),
            "cpu_flow": get_cell_payload(ws, row_lookup.get(SEMANTIC_ROW_LABELS["cpu_flow"]), col_idx),
            "gcd_flow": get_cell_payload(ws, row_lookup.get(SEMANTIC_ROW_LABELS["gcd_flow"]), col_idx),
            "hub_flow": get_cell_payload(ws, row_lookup.get(SEMANTIC_ROW_LABELS["hub_flow"]), col_idx),
            "pcd_flow": get_cell_payload(ws, row_lookup.get(SEMANTIC_ROW_LABELS["pcd_flow"]), col_idx),
        }

        for dielet in DIELET_KEYS:
            prefix = dielet.lower()
            for meta_key in DIELET_METADATA_KEYS:
                field_name = f"{prefix}_{meta_key}"
                slot[field_name] = get_cell_payload(ws, row_lookup.get(SEMANTIC_ROW_LABELS[field_name]), col_idx)

        if fullbp_row_labels:
            slot["fullbp_rows"] = {
                row_label: get_cell_payload(ws, row_lookup[row_label], col_idx)
                for row_label in fullbp_row_labels
            }

        if ws.title == "SimplerBPView" and fullbp_index is not None:
            flow_key = normalize_text(all_flow["value"])
            fullbp_slot = fullbp_index.get(flow_key)
            if fullbp_slot is not None:
                slot["resolved_from_fullbp"] = {
                    "slot_col": fullbp_slot["slot_col"],
                    "slot_col_index": fullbp_slot["slot_col_index"],
                    "tos_exec_flow": fullbp_slot.get("tos_exec_flow"),
                    "tp_exec_flow": fullbp_slot.get("tp_exec_flow"),
                    "top_flow": fullbp_slot.get("top_flow"),
                    "top_flow_type": fullbp_slot.get("top_flow_type"),
                    "top_port_range": fullbp_slot.get("top_port_range"),
                    "top_dielet_type": fullbp_slot.get("top_dielet_type"),
                    "all_dielet_type": fullbp_slot.get("all_dielet_type"),
                    "all_scope": fullbp_slot.get("all_scope"),
                    "all_port_range": fullbp_slot.get("all_port_range"),
                    "all_flow_type": fullbp_slot.get("all_flow_type"),
                }
                for dielet in DIELET_KEYS:
                    prefix = dielet.lower()
                    for meta_key in DIELET_METADATA_KEYS:
                        field_name = f"{prefix}_{meta_key}"
                        slot["resolved_from_fullbp"][field_name] = fullbp_slot.get(field_name)

        slots.append(slot)

    return slots


def build_slot_search_text(slot):
    """Flatten slot fields into a list of searchable strings."""
    values = []
    for field in [
        "tos_exec_flow", "tp_exec_flow", "top_flow", "top_flow_type", "top_port_range", "top_dielet_type",
        "all_flow", "all_dielet_type", "all_scope", "all_port_range", "all_flow_type",
        "cpu_flow", "gcd_flow", "hub_flow", "pcd_flow"
    ]:
        payload = slot.get(field)
        if payload and payload.get("value") is not None:
            values.append(normalize_text(payload["value"]))

    for dielet in DIELET_KEYS:
        prefix = dielet.lower()
        for meta_key in DIELET_METADATA_KEYS:
            field_name = f"{prefix}_{meta_key}"
            payload = slot.get(field_name)
            if payload and payload.get("value") is not None:
                values.append(normalize_text(payload["value"]))

    resolved = slot.get("resolved_from_fullbp")
    if resolved:
        for resolved_field in TOP_LEVEL_FIELD_KEYS + ["all_dielet_type", "all_scope", "all_port_range", "all_flow_type"]:
            resolved_payload = resolved.get(resolved_field)
            if resolved_payload and resolved_payload.get("value") is not None:
                values.append(normalize_text(resolved_payload["value"]))
        for dielet in DIELET_KEYS:
            prefix = dielet.lower()
            for meta_key in DIELET_METADATA_KEYS:
                field_name = f"{prefix}_{meta_key}"
                resolved_payload = resolved.get(field_name)
                if resolved_payload and resolved_payload.get("value") is not None:
                    values.append(normalize_text(resolved_payload["value"]))

    return values


def build_slot_indexes(slots):
    """Index slots by normalized FLOWDEF:ALL:Flow value."""
    index = {}
    for slot in slots:
        flow_value = normalize_text(slot.get("all_flow", {}).get("value"))
        if flow_value:
            index.setdefault(flow_value, []).append(slot)
    return index


def slot_dielet_signature(slot):
    """Return the ordered dielet signature for a slot."""
    signature = []
    for dielet in DIELET_KEYS:
        payload = slot.get(DIELET_FIELD_MAP[dielet])
        if payload and normalize_text(payload.get("value")):
            signature.append(dielet)
    return tuple(signature)


def slot_metadata_snapshot(slot):
    """Return comparable metadata values for ambiguity detection."""
    snapshot = {}
    for field in ["all_dielet_type", "all_scope", "all_port_range", "all_flow_type"]:
        payload = slot.get(field)
        snapshot[field] = normalize_text(payload.get("value")) if payload else ""
    return snapshot


def column_style_signature(ws, col_idx):
    """Build a comparable signature for styles in a column."""
    signature = []
    for row_idx in range(1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        signature.append((row_idx, cell.style_id))
    return tuple(signature)


def copy_column_style(ws, source_col_idx, target_col_idx):
    """Copy style-only information from one column to another."""
    for row_idx in range(1, ws.max_row + 1):
        source_cell = ws.cell(row=row_idx, column=source_col_idx)
        target_cell = ws.cell(row=row_idx, column=target_col_idx)
        if source_cell.has_style:
            target_cell._style = copy.copy(source_cell._style)
        if source_cell.font:
            target_cell.font = copy.copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy.copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy.copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy.copy(source_cell.alignment)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.protection:
            target_cell.protection = copy.copy(source_cell.protection)

    source_letter = get_column_letter(source_col_idx)
    target_letter = get_column_letter(target_col_idx)
    if source_letter in ws.column_dimensions:
        ws.column_dimensions[target_letter] = copy.copy(ws.column_dimensions[source_letter])
        ws.column_dimensions[target_letter].index = target_letter


def snapshot_column_dimensions(ws):
    """Capture explicit worksheet column dimension metadata keyed by column index."""
    snapshot = {}
    for column_letter in list(ws.column_dimensions.keys()):
        try:
            column_idx = column_index_from_string(column_letter)
        except Exception:
            continue
        snapshot[column_idx] = copy.copy(ws.column_dimensions[column_letter])
    return snapshot


def restore_shifted_column_dimensions(ws, dimension_snapshot, insert_col, amount=1):
    """Restore explicit column dimensions after columns are inserted."""
    for column_letter in list(ws.column_dimensions.keys()):
        try:
            column_index_from_string(column_letter)
        except Exception:
            continue
        del ws.column_dimensions[column_letter]

    for original_idx, original_dimension in sorted(dimension_snapshot.items()):
        new_idx = original_idx + amount if original_idx >= insert_col else original_idx
        new_letter = get_column_letter(new_idx)
        ws.column_dimensions[new_letter] = copy.copy(original_dimension)
        ws.column_dimensions[new_letter].index = new_letter


def choose_fullbp_style_donor(context, anchor_slot, insert_col, mode):
    """Choose a nearby FullBP style donor that stays in the local top-flow family."""
    req_flow_type = requested_flow_type(mode)
    top_level_anchor_fields = [
        field_name for field_name in TOP_LEVEL_FIELD_KEYS
        if normalize_text(anchor_slot.get(field_name, {}).get("value") if anchor_slot.get(field_name) else None)
    ]
    candidates = []

    for slot in context["fullbp_slots"]:
        slot_flow_type = normalize_text(slot.get("all_flow_type", {}).get("value"))
        if slot_flow_type != req_flow_type:
            continue

        top_mismatch_count = 0
        for field_name in top_level_anchor_fields:
            anchor_value = normalize_text(anchor_slot.get(field_name, {}).get("value"))
            candidate_value = normalize_text(slot.get(field_name, {}).get("value"))
            if candidate_value != anchor_value:
                top_mismatch_count += 1

        distance = abs(slot["slot_col_index"] - insert_col)
        candidates.append((top_mismatch_count, distance, slot["slot_col_index"], slot))

    if not candidates:
        return anchor_slot

    candidates.sort(key=lambda item: item[:3])
    return candidates[0][3]


def parse_dielet_flows(raw_values):
    """Parse repeated DIELET=FLOW arguments into a normalized mapping."""
    flows = {}
    duplicates = []
    for raw_value in raw_values or []:
        if "=" not in raw_value:
            raise ValueError(f"Invalid --dielet-flow '{raw_value}'. Expected DIELET=FLOWNAME.")
        dielet, flow_name = raw_value.split("=", 1)
        dielet = dielet.strip().upper()
        flow_name = flow_name.strip()
        if dielet not in DIELET_KEYS:
            raise ValueError(f"Invalid dielet '{dielet}'. Use one of: {', '.join(DIELET_KEYS)}")
        if not flow_name:
            raise ValueError(f"Empty flow name for dielet '{dielet}'.")
        if dielet in flows:
            duplicates.append(dielet)
        flows[dielet] = flow_name

    if duplicates:
        raise ValueError(
            f"Duplicate dielet flow definitions for: {', '.join(sorted(set(duplicates)))}"
        )

    return flows


def requested_flow_type(mode):
    """Convert serial/parallel input into the workbook flow type value."""
    return PARALLEL_FLOW_TYPE if mode == "parallel" else SERIAL_FLOW_TYPE


def flow_name_list(raw_values):
    """Return normalized flow names preserving order."""
    return [normalize_text(value) for value in raw_values if normalize_text(value)]


def find_named_slots(slot_index, flow_name):
    """Return slots matching a FLOWDEF:ALL:Flow value."""
    return slot_index.get(normalize_text(flow_name), [])


def serialize_slot_brief(slot):
    """Return a brief JSON-friendly slot summary."""
    return {
        "sheet": slot.get("sheet"),
        "slot_col": slot.get("slot_col"),
        "slot_col_index": slot.get("slot_col_index"),
        "tos_exec_flow": slot.get("tos_exec_flow", {}).get("value") if slot.get("tos_exec_flow") else None,
        "tp_exec_flow": slot.get("tp_exec_flow", {}).get("value") if slot.get("tp_exec_flow") else None,
        "top_flow": slot.get("top_flow", {}).get("value") if slot.get("top_flow") else None,
        "all_flow": slot.get("all_flow", {}).get("value") if slot.get("all_flow") else None,
        "all_flow_type": slot.get("all_flow_type", {}).get("value") if slot.get("all_flow_type") else None,
        "signature": list(slot_dielet_signature(slot)),
    }


def build_workbook_context(wb):
    """Build reusable workbook slot context for planning and updates."""
    fullbp_ws = get_sheet(wb, "FullBP")
    fullbp_row_lookup = build_row_lookup(fullbp_ws)
    fullbp_row_labels = list_fullbp_row_labels(fullbp_ws)
    fullbp_slots = build_slot_view(fullbp_ws, include_fullbp_rows=True)
    fullbp_index = build_slot_indexes(fullbp_slots)

    simpler_ws = get_sheet(wb, "SimplerBPView")
    simpler_slots = build_slot_view(simpler_ws, fullbp_index={
        normalize_text(slot["all_flow"]["value"]): slot
        for slot in fullbp_slots
        if slot.get("all_flow") and normalize_text(slot["all_flow"]["value"])
    })
    simpler_index = build_slot_indexes(simpler_slots)

    return {
        "fullbp_ws": fullbp_ws,
        "fullbp_row_lookup": fullbp_row_lookup,
        "fullbp_row_labels": fullbp_row_labels,
        "fullbp_slots": fullbp_slots,
        "fullbp_index": fullbp_index,
        "simpler_ws": simpler_ws,
        "simpler_slots": simpler_slots,
        "simpler_index": simpler_index,
    }


def choose_fullbp_template(context, anchor_slot, mode, dielet_flows):
    """Choose the best FullBP template slot for metadata/style inference."""
    req_signature = tuple(dielet for dielet in DIELET_KEYS if dielet in dielet_flows)
    req_flow_type = requested_flow_type(mode)
    candidates = []
    top_level_anchor_fields = [
        field_name for field_name in TOP_LEVEL_FIELD_KEYS
        if normalize_text(anchor_slot.get(field_name, {}).get("value") if anchor_slot.get(field_name) else None)
    ]

    for slot in context["fullbp_slots"]:
        slot_flow_type = normalize_text(slot.get("all_flow_type", {}).get("value"))
        if slot_flow_type != req_flow_type:
            continue

        top_mismatch_count = 0
        top_missing_count = 0
        for field_name in top_level_anchor_fields:
            anchor_value = normalize_text(anchor_slot.get(field_name, {}).get("value"))
            candidate_value = normalize_text(slot.get(field_name, {}).get("value"))
            if candidate_value:
                if candidate_value != anchor_value:
                    top_mismatch_count += 1
            else:
                top_missing_count += 1

        slot_signature = slot_dielet_signature(slot)
        if slot_signature == req_signature:
            signature_rank = 0
        elif set(slot_signature) == set(req_signature):
            signature_rank = 1
        elif req_signature and set(req_signature).issubset(set(slot_signature)):
            signature_rank = 2
        else:
            signature_rank = 3

        distance = abs(slot["slot_col_index"] - anchor_slot["slot_col_index"])
        candidates.append((
            top_mismatch_count,
            top_missing_count,
            signature_rank,
            distance,
            slot["slot_col_index"],
            slot,
        ))

    if not candidates:
        return None, ["No FullBP template slot found with matching flow type."], []

    candidates.sort(key=lambda item: item[:5])
    best_mismatch, best_missing, best_rank, best_distance, _, best_slot = candidates[0]
    best_group = [
        item[5] for item in candidates
        if item[0] == best_mismatch and item[1] == best_missing and item[2] == best_rank and item[3] == best_distance
    ]

    ambiguities = []
    style_signatures = [
        column_style_signature(context["fullbp_ws"], slot["slot_col_index"])
        for slot in best_group
    ]
    if len(best_group) > 1 and len(set(style_signatures)) > 1:
        ambiguities.append("FullBP style template is ambiguous across equally good candidate slots.")

    return best_slot, ambiguities, best_group, [serialize_slot_brief(slot) for slot in best_group]


def choose_simpler_template(context, anchor_slot, fullbp_template_slot):
    """Choose the SimplerBP style template corresponding to the selected FullBP template."""
    if fullbp_template_slot is None:
        return anchor_slot, [
            "SimplerBP style template fell back to the anchor slot because no FullBP style donor was selected."
        ]
    template_flow_name = normalize_text(fullbp_template_slot.get("all_flow", {}).get("value"))
    matches = find_named_slots(context["simpler_index"], template_flow_name)
    if len(matches) == 1:
        return matches[0], []
    if len(matches) > 1:
        return anchor_slot, [
            "SimplerBP style template matched multiple slots; falling back to anchor slot style."
        ]
    return anchor_slot, []


def choose_style_templates(context, simpler_anchor, fullbp_anchor, insert_col, mode,
                           resolved_scope=None, resolved_dielet_type=None):
    """Choose validated style donor slots used during insert writeback."""
    notes = []
    issues = []
    req_flow_type = requested_flow_type(mode)
    resolved_scope_text = normalize_text(resolved_scope).upper()
    resolved_dielet_text = normalize_text(resolved_dielet_type).upper()
    top_level_anchor_fields = [
        field_name for field_name in TOP_LEVEL_FIELD_KEYS
        if normalize_text(fullbp_anchor.get(field_name, {}).get("value") if fullbp_anchor.get(field_name) else None)
    ]

    candidates = []
    for slot in context["fullbp_slots"]:
        slot_flow_type = normalize_text(slot.get("all_flow_type", {}).get("value"))
        if slot_flow_type != req_flow_type:
            continue

        slot_scope = normalize_text(slot.get("all_scope", {}).get("value")).upper()
        if resolved_scope_text and slot_scope != resolved_scope_text:
            continue

        slot_dielet_type = normalize_text(slot.get("all_dielet_type", {}).get("value")).upper()
        if resolved_dielet_text and slot_dielet_type != resolved_dielet_text:
            continue

        top_mismatch_count = 0
        for field_name in top_level_anchor_fields:
            anchor_value = normalize_text(fullbp_anchor.get(field_name, {}).get("value"))
            candidate_value = normalize_text(slot.get(field_name, {}).get("value"))
            if candidate_value != anchor_value:
                top_mismatch_count += 1

        distance = abs(slot["slot_col_index"] - insert_col)
        candidates.append((top_mismatch_count, distance, slot["slot_col_index"], slot))

    if not candidates:
        issues.append(
            "No FullBP style donor matched the resolved scope/dielet-type family."
        )
        fullbp_style_slot = None
    else:
        candidates.sort(key=lambda item: item[:3])
        fullbp_style_slot = candidates[0][3]
        notes.append("Using a scope-validated FullBP style donor.")

    simpler_style_slot, simpler_notes = choose_simpler_template(context, simpler_anchor, fullbp_style_slot)
    notes.extend(simpler_notes)

    return {
        "fullbp": fullbp_style_slot,
        "simpler": simpler_style_slot,
        "notes": notes,
        "issues": issues,
    }


def detect_dielet_tokens(flow_name):
    """Return all dielet tokens present in the flow name."""
    flow_upper = normalize_text(flow_name).upper()
    return [dielet for dielet in DIELET_KEYS if dielet in flow_upper]


def slot_row_payload(slot, row_label):
    """Return a payload for a row label from a FullBP slot or a semantic slot field."""
    fullbp_rows = slot.get("fullbp_rows")
    if fullbp_rows is not None:
        return fullbp_rows.get(row_label)
    field_name = SEMANTIC_FIELD_BY_ROW_LABEL.get(row_label)
    if field_name:
        return slot.get(field_name)
    return None


def collect_template_row_values(template_candidates, row_label):
    """Collect non-empty candidate values for a given FullBP row label."""
    values = {}
    for candidate in template_candidates:
        payload = slot_row_payload(candidate, row_label)
        if payload is None or payload.get("value") is None:
            continue
        normalized = normalize_text(payload["value"])
        if not normalized:
            continue
        entry = values.setdefault(normalized, {
            "value": payload["value"],
            "sources": [],
        })
        entry["sources"].append(serialize_slot_brief(candidate))
    return values


def extract_flowdef_dielet(row_label):
    """Return (dielet, row_kind) for FLOWDEF dielet rows, or (None, None)."""
    parts = row_label.split(":")
    if len(parts) == 3 and parts[0] == "FLOWDEF" and parts[1] in DIELET_KEYS:
        return parts[1], parts[2]
    return None, None


def build_fullbp_row_plan_entry(row_label, row_index, template_candidates, selected_template, field_name,
                                explicit_value=None, blank=False, blank_reason=None):
    """Build a row-level plan entry for a single FullBP label."""
    entry = {
        "row_label": row_label,
        "row": row_index,
        "field": field_name,
    }

    if explicit_value is not None:
        entry["status"] = "explicit"
        entry["value"] = explicit_value
        return entry

    if blank:
        entry["status"] = "blank"
        if blank_reason:
            entry["reason"] = blank_reason
        return entry

    candidate_values = collect_template_row_values(template_candidates, row_label)
    selected_payload = slot_row_payload(selected_template, row_label)
    selected_value = None
    if selected_payload is not None:
        selected_value = selected_payload.get("value")

    if len(candidate_values) == 1:
        entry["status"] = "inferred"
        entry["value"] = next(iter(candidate_values.values()))["value"]
        entry["sources"] = next(iter(candidate_values.values()))["sources"]
        return entry

    if len(candidate_values) > 1:
        if len(template_candidates) == 1 and selected_value is not None:
            entry["status"] = "copied_from_template"
            entry["value"] = selected_value
            entry["source"] = serialize_slot_brief(selected_template)
            return entry
        entry["status"] = "ambiguous"
        entry["candidates"] = list(candidate_values.values())
        return entry

    if selected_value is not None:
        entry["status"] = "copied_from_template"
        entry["value"] = selected_value
        entry["source"] = serialize_slot_brief(selected_template)
        return entry

    entry["status"] = "blank"
    return entry


def summarize_fullbp_row_plan(row_plan):
    """Build legacy summaries and ambiguity lists from the row-complete plan."""
    inferred = {}
    inferred_dielet = {}
    ambiguities = []
    for entry in row_plan:
        field_name = entry.get("field")
        status = entry.get("status")
        if status == "ambiguous":
            ambiguity_entry = {
                "row_label": entry["row_label"],
                "candidates": [candidate["value"] for candidate in entry.get("candidates", [])],
            }
            if field_name:
                ambiguity_entry["field"] = field_name
            ambiguities.append(ambiguity_entry)
            continue

        if field_name and status in PLAN_VALUE_STATUSES and entry.get("value") not in (None, ""):
            if field_name.startswith(("cpu_", "gcd_", "hub_", "pcd_")) and field_name not in DIELET_FIELD_MAP.values():
                inferred_dielet[field_name] = entry["value"]
            elif field_name not in ["all_flow"] + list(DIELET_FIELD_MAP.values()):
                inferred[field_name] = entry["value"]

    return inferred, inferred_dielet, ambiguities


def detect_dielet_hint(flow_name):
    """Return a single dielet token from a flow name when the signal is unambiguous."""
    flow_upper = normalize_text(flow_name).upper()
    detected = [dielet for dielet in DIELET_KEYS if dielet in flow_upper]
    if len(detected) == 1:
        return detected[0]
    return None


def detect_scope_hint(flow_name):
    """Return a coarse scope hint derived from the flow name."""
    flow_upper = normalize_text(flow_name).upper()
    has_pkg = "PKG" in flow_upper
    has_ip = "IP" in flow_upper
    if has_pkg and has_ip:
        return None
    if has_pkg:
        return "PKG"
    if has_ip:
        return "IP"
    return None


def scope_matches_hint(scope_value, scope_hint):
    """Return True when a workbook scope value satisfies a coarse scope hint."""
    scope_text = normalize_text(scope_value).upper()
    if not scope_hint:
        return True
    if scope_hint == "PKG":
        return scope_text == "PKG"
    if scope_hint == "IP":
        return scope_text.startswith("IP")
    return False


def classify_flow_owner(flow_name, name_inference):
    """Classify whether a flow most likely belongs to Shared/common or the active dielet repo."""
    dielet_tokens = detect_dielet_tokens(flow_name)
    flow_upper = normalize_text(flow_name).upper()
    has_pkg = "PKG" in flow_upper
    has_ip = "IP" in flow_upper

    if has_pkg and has_ip:
        return {
            "status": "ambiguous",
            "owner": None,
            "reason": "Flow name contains both PKG and IP tokens.",
            "requires_confirmation": False,
        }

    if len(dielet_tokens) > 1:
        return {
            "status": "ambiguous",
            "owner": None,
            "reason": f"Flow name contains multiple dielet tokens: {', '.join(dielet_tokens)}.",
            "requires_confirmation": False,
        }

    if has_ip:
        return {
            "status": "classified",
            "owner": "dielet",
            "reason": "IP-named subflows are dielet-local.",
            "requires_confirmation": True,
        }

    if has_pkg and dielet_tokens:
        return {
            "status": "classified",
            "owner": "dielet",
            "reason": f"PKG flow name includes dielet token {dielet_tokens[0]}.",
            "requires_confirmation": True,
        }

    if has_pkg:
        return {
            "status": "classified",
            "owner": "shared",
            "reason": "PKG flow name has no dielet token, so it is treated as Shared/common until confirmed.",
            "requires_confirmation": True,
        }

    return {
        "status": "ambiguous",
        "owner": None,
        "reason": "Flow name does not identify PKG/IP ownership clearly enough to classify automatically.",
        "requires_confirmation": False,
    }


def row_plan_value(row_plan, field_name):
    """Return the resolved value for a planned row field when available."""
    for entry in row_plan:
        if entry.get("field") != field_name:
            continue
        if entry.get("status") in PLAN_VALUE_STATUSES and entry.get("value") not in (None, ""):
            return entry.get("value")
    return None


def choose_name_inference_template(context, anchor_slot, mode, new_flow):
    """Choose a narrower FullBP template group using subflow-name hints plus workbook evidence."""
    dielet_hint = detect_dielet_hint(new_flow)
    scope_hint = detect_scope_hint(new_flow)
    if dielet_hint is None and scope_hint is None:
        return {
            "dielet_hint": None,
            "scope_hint": None,
            "expected_top_flow": normalize_text(
                anchor_slot.get("top_flow", {}).get("value") if anchor_slot.get("top_flow") else None
            ),
            "template": None,
            "candidates": [],
        }

    requested_type = requested_flow_type(mode)
    expected_top_flow = normalize_text(
        anchor_slot.get("top_flow", {}).get("value") if anchor_slot.get("top_flow") else None
    )
    flow_upper = normalize_text(new_flow).upper()
    name_tokens = {
        token for token in [expected_top_flow, dielet_hint, scope_hint]
        if token and token in flow_upper
    }

    def collect_candidates(require_top_flow):
        candidates = []
        for slot in context["fullbp_slots"]:
            if normalize_text(slot.get("all_flow_type", {}).get("value")) != requested_type:
                continue

            slot_top_flow = normalize_text(slot.get("top_flow", {}).get("value"))
            if require_top_flow and expected_top_flow and slot_top_flow != expected_top_flow:
                continue

            slot_dielet = normalize_text(slot.get("all_dielet_type", {}).get("value"))
            if dielet_hint and slot_dielet != dielet_hint:
                continue

            slot_scope = normalize_text(slot.get("all_scope", {}).get("value"))
            if scope_hint and not scope_matches_hint(slot_scope, scope_hint):
                continue

            slot_name = normalize_text(slot.get("all_flow", {}).get("value")).upper()
            overlap = sum(1 for token in name_tokens if token in slot_name)
            distance = abs(slot["slot_col_index"] - anchor_slot["slot_col_index"])
            candidates.append((-overlap, distance, slot["slot_col_index"], slot))

        candidates.sort(key=lambda item: item[:3])
        return candidates

    ranked_candidates = collect_candidates(require_top_flow=True)
    if not ranked_candidates:
        ranked_candidates = collect_candidates(require_top_flow=False)
    if not ranked_candidates:
        return {
            "dielet_hint": dielet_hint,
            "scope_hint": scope_hint,
            "expected_top_flow": expected_top_flow,
            "template": None,
            "candidates": [],
        }

    best_overlap = ranked_candidates[0][0]
    best_group = [item[3] for item in ranked_candidates if item[0] == best_overlap]
    selected_template = sorted(
        best_group,
        key=lambda slot: (abs(slot["slot_col_index"] - anchor_slot["slot_col_index"]), slot["slot_col_index"]),
    )[0]
    return {
        "dielet_hint": dielet_hint,
        "scope_hint": scope_hint,
        "expected_top_flow": expected_top_flow,
        "template": selected_template,
        "candidates": best_group,
    }


def plan_insert_slot(args, wb):
    """Build an insert-slot plan without mutating the workbook."""
    context = build_workbook_context(wb)
    try:
        dielet_flows = parse_dielet_flows(args.dielet_flow)
    except ValueError as exc:
        return {"status": "blocked", "blocking_issues": [str(exc)]}
    requested_signature = [dielet for dielet in DIELET_KEYS if dielet in dielet_flows]

    simpler_matches = find_named_slots(context["simpler_index"], args.anchor_flow)
    if not simpler_matches:
        return {"status": "blocked", "blocking_issues": [
            f"Anchor flow '{args.anchor_flow}' was not found in SimplerBPView."
        ]}
    if len(simpler_matches) > 1:
        return {"status": "blocked", "blocking_issues": [
            f"Anchor flow '{args.anchor_flow}' is ambiguous in SimplerBPView.",
        ], "anchor_candidates": [serialize_slot_brief(slot) for slot in simpler_matches]}

    simpler_anchor = simpler_matches[0]
    fullbp_anchor_matches = find_named_slots(context["fullbp_index"], args.anchor_flow)
    if len(fullbp_anchor_matches) != 1:
        return {"status": "blocked", "blocking_issues": [
            f"Anchor flow '{args.anchor_flow}' could not be mapped uniquely into FullBP."
        ], "anchor_candidates": [serialize_slot_brief(slot) for slot in fullbp_anchor_matches]}

    fullbp_anchor = fullbp_anchor_matches[0]
    fullbp_template, template_ambiguities, template_group_slots, template_group = choose_fullbp_template(
        context, fullbp_anchor, args.mode, dielet_flows
    )
    if fullbp_template is None:
        return {
            "status": "blocked",
            "blocking_issues": ["No suitable FullBP template slot could be inferred."],
        }

    simpler_template, simpler_template_notes = choose_simpler_template(
        context, simpler_anchor, fullbp_template
    )

    blocking_issues = []
    if template_ambiguities:
        blocking_issues.extend(template_ambiguities)

    insert_offset = 0 if args.position == "before" else 1
    name_inference = choose_name_inference_template(context, fullbp_anchor, args.mode, args.new_flow)
    owner_classification = classify_flow_owner(args.new_flow, name_inference)
    if name_inference.get("dielet_hint") and args.dielet_type is not None:
        explicit_dielet_type = normalize_text(args.dielet_type).upper()
        if explicit_dielet_type and explicit_dielet_type != name_inference["dielet_hint"]:
            blocking_issues.append(
                f"Flow name '{args.new_flow}' implies dielet '{name_inference['dielet_hint']}' but --dielet-type was '{args.dielet_type}'."
            )
    if name_inference.get("scope_hint") and args.scope is not None:
        if not scope_matches_hint(args.scope, name_inference["scope_hint"]):
            blocking_issues.append(
                f"Flow name '{args.new_flow}' implies scope family '{name_inference['scope_hint']}' but --scope was '{args.scope}'."
            )
    if owner_classification.get("status") == "ambiguous":
        if getattr(args, "confirm_owner", None):
            owner_classification = {
                "status": "classified",
                "owner": args.confirm_owner,
                "reason": f"Owner confirmed by user via --confirm-owner {args.confirm_owner}.",
                "requires_confirmation": False,
            }
        else:
            blocking_issues.append(owner_classification["reason"])

    fullbp_row_plan = []
    for row_label in context["fullbp_row_labels"]:
        row_index = context["fullbp_row_lookup"][row_label]
        field_name = SEMANTIC_FIELD_BY_ROW_LABEL.get(row_label)
        explicit_value = None
        blank = False
        blank_reason = None
        row_template_candidates = template_group_slots
        row_selected_template = fullbp_template

        if row_label in {
            SEMANTIC_ROW_LABELS["all_dielet_type"],
            SEMANTIC_ROW_LABELS["all_scope"],
            SEMANTIC_ROW_LABELS["all_port_range"],
        } and name_inference.get("candidates"):
            row_template_candidates = name_inference["candidates"]
            row_selected_template = name_inference["template"]

        if row_label == SEMANTIC_ROW_LABELS["all_flow"]:
            explicit_value = args.new_flow
        elif row_label == SEMANTIC_ROW_LABELS["all_flow_type"]:
            explicit_value = requested_flow_type(args.mode)
        elif row_label == SEMANTIC_ROW_LABELS["all_dielet_type"] and args.dielet_type is not None:
            explicit_value = args.dielet_type
        elif row_label == SEMANTIC_ROW_LABELS["all_scope"] and args.scope is not None:
            explicit_value = args.scope
        elif row_label == SEMANTIC_ROW_LABELS["all_port_range"] and args.port_range is not None:
            explicit_value = args.port_range
        else:
            dielet, row_kind = extract_flowdef_dielet(row_label)
            if dielet is not None:
                if row_kind == "Flow":
                    if dielet in dielet_flows:
                        explicit_value = dielet_flows[dielet]
                    else:
                        blank = True
                        blank_reason = f"{dielet} does not participate in this slot."
                elif dielet not in dielet_flows:
                    blank = True
                    blank_reason = f"{dielet} does not participate in this slot."

        fullbp_row_plan.append(build_fullbp_row_plan_entry(
            row_label,
            row_index,
            row_template_candidates,
            row_selected_template,
            field_name,
            explicit_value=explicit_value,
            blank=blank,
            blank_reason=blank_reason,
        ))

    inferred, inferred_dielet, ambiguities = summarize_fullbp_row_plan(fullbp_row_plan)
    resolved_scope = row_plan_value(fullbp_row_plan, "all_scope")
    resolved_dielet_type = row_plan_value(fullbp_row_plan, "all_dielet_type")
    style_templates = choose_style_templates(
        context,
        simpler_anchor,
        fullbp_anchor,
        fullbp_anchor["slot_col_index"] + insert_offset,
        args.mode,
        resolved_scope=resolved_scope,
        resolved_dielet_type=resolved_dielet_type,
    )
    if style_templates.get("issues"):
        blocking_issues.extend(style_templates["issues"])

    plan = {
        "status": "ready" if not blocking_issues else "blocked",
        "blocking_issues": blocking_issues,
        "anchor": {
            "simpler": serialize_slot_brief(simpler_anchor),
            "fullbp": serialize_slot_brief(fullbp_anchor),
        },
        "request": {
            "new_flow": args.new_flow,
            "position": args.position,
            "mode": args.mode,
            "requested_signature": requested_signature,
            "dielet_flows": dielet_flows,
        },
        "templates": {
            "fullbp": serialize_slot_brief(fullbp_template),
            "fullbp_candidates": template_group,
            "simpler": serialize_slot_brief(simpler_template),
            "notes": simpler_template_notes,
        },
        "style_templates": {
            "fullbp": serialize_slot_brief(style_templates["fullbp"]),
            "simpler": serialize_slot_brief(style_templates["simpler"]),
            "notes": style_templates["notes"],
        },
        "name_inference": {
            "dielet_hint": name_inference.get("dielet_hint"),
            "scope_hint": name_inference.get("scope_hint"),
            "expected_top_flow": name_inference.get("expected_top_flow"),
            "fullbp": serialize_slot_brief(name_inference["template"]) if name_inference.get("template") else None,
            "candidate_count": len(name_inference.get("candidates", [])),
        },
        "owner_classification": owner_classification,
        "confirmation_required": owner_classification["owner"] if owner_classification.get("requires_confirmation") else None,
        "targets": {
            "simpler_insert_col": simpler_anchor["slot_col_index"] + insert_offset,
            "fullbp_insert_col": fullbp_anchor["slot_col_index"] + insert_offset,
        },
        "inferred": inferred,
        "inferred_dielet": inferred_dielet,
        "ambiguities": ambiguities,
        "fullbp_row_plan": fullbp_row_plan,
    }
    return plan


def select_occurrence(slots, occurrence):
    """Select a 1-based occurrence from a list of slots sorted by column index."""
    sorted_slots = sorted(slots, key=lambda slot: slot["slot_col_index"])
    if occurrence is None:
        if len(sorted_slots) == 1:
            return sorted_slots[0], []
        return None, [serialize_slot_brief(slot) for slot in sorted_slots]

    if occurrence < 1 or occurrence > len(sorted_slots):
        return None, [serialize_slot_brief(slot) for slot in sorted_slots]

    return sorted_slots[occurrence - 1], []


def plan_delete_slot(args, wb):
    """Build a delete-slot plan without mutating the workbook."""
    context = build_workbook_context(wb)

    simpler_matches = find_named_slots(context["simpler_index"], args.target_flow)
    if not simpler_matches:
        return {
            "status": "blocked",
            "blocking_issues": [f"Target flow '{args.target_flow}' was not found in SimplerBPView."],
        }

    simpler_target, simpler_candidates = select_occurrence(simpler_matches, args.occurrence)
    if simpler_target is None:
        message = f"Target flow '{args.target_flow}' is ambiguous in SimplerBPView."
        if args.occurrence is not None:
            message = (
                f"Occurrence {args.occurrence} is out of range for target flow '{args.target_flow}' "
                "in SimplerBPView."
            )
        return {
            "status": "blocked",
            "blocking_issues": [message],
            "target_candidates": simpler_candidates,
        }

    fullbp_matches = find_named_slots(context["fullbp_index"], args.target_flow)
    if not fullbp_matches:
        return {
            "status": "blocked",
            "blocking_issues": [f"Target flow '{args.target_flow}' was not found in FullBP."],
        }

    fullbp_target, fullbp_candidates = select_occurrence(fullbp_matches, args.occurrence)
    if fullbp_target is None:
        message = f"Target flow '{args.target_flow}' is ambiguous in FullBP."
        if args.occurrence is not None:
            message = (
                f"Occurrence {args.occurrence} is out of range for target flow '{args.target_flow}' "
                "in FullBP."
            )
        return {
            "status": "blocked",
            "blocking_issues": [message],
            "target_candidates": fullbp_candidates,
        }

    return {
        "status": "ready",
        "blocking_issues": [],
        "request": {
            "target_flow": args.target_flow,
            "occurrence": args.occurrence,
        },
        "targets": {
            "simpler": serialize_slot_brief(simpler_target),
            "fullbp": serialize_slot_brief(fullbp_target),
        },
    }


def backup_file(path):
    """Create a .bak copy of the file. Returns the backup path."""
    bak_path = path.with_suffix(path.suffix + ".bak")
    shutil.copy2(str(path), str(bak_path))
    return bak_path


# ---------------------------------------------------------------------------
# Command implementations
# ---------------------------------------------------------------------------

def cmd_read(args):
    """Dump one or all sheets as JSON with cell values and fill colors."""
    excel_path = get_excel_path()
    wb = load_workbook_readonly(excel_path)
    sheets_to_dump = [args.sheet] if args.sheet else [
        sheet_name for sheet_name in ALLOWED_SHEETS if sheet_name in wb.sheetnames
    ]

    fullbp_slots = []
    fullbp_index = {}
    if "FullBP" in wb.sheetnames:
        fullbp_slots = build_slot_view(wb["FullBP"])
        fullbp_index = {
            normalize_text(slot["all_flow"]["value"]): slot
            for slot in fullbp_slots
            if slot.get("all_flow") and normalize_text(slot["all_flow"]["value"])
        }

    result = {
        "file": str(excel_path),
        "all_sheets": wb.sheetnames,
        "allowed_sheets": sheets_to_dump,
        "sheets": {},
        "slot_view": {}
    }
    for sname in sheets_to_dump:
        ws = get_sheet(wb, sname)
        result["sheets"][sname] = sheet_to_rows(ws)
        if sname == "FullBP":
            result["slot_view"][sname] = fullbp_slots
        else:
            result["slot_view"][sname] = build_slot_view(ws, fullbp_index=fullbp_index)

    print(json.dumps(result, default=str, indent=2))


def cmd_query(args):
    """Search all rows for cells containing the query text (case-insensitive)."""
    wb = load_workbook_readonly(get_excel_path())
    search_text = args.query.lower()
    sheets_to_search = [args.sheet] if args.sheet else [
        sheet_name for sheet_name in ALLOWED_SHEETS if sheet_name in wb.sheetnames
    ]

    fullbp_slots = []
    fullbp_index = {}
    if "FullBP" in wb.sheetnames:
        fullbp_slots = build_slot_view(wb["FullBP"])
        fullbp_index = {
            normalize_text(slot["all_flow"]["value"]): slot
            for slot in fullbp_slots
            if slot.get("all_flow") and normalize_text(slot["all_flow"]["value"])
        }

    matches = {}
    slot_matches = {}
    for sname in sheets_to_search:
        ws = get_sheet(wb, sname)
        sheet_matches = []
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            row_has_match = any(
                search_text in str(cell.value).lower()
                for cell in row
                if cell.value is not None
            )
            if row_has_match:
                cells = []
                for cell in row:
                    cells.append({
                        "col": get_column_letter(cell.column),
                        "col_index": cell.column,
                        "value": cell.value,
                        "color": get_cell_color(cell)
                    })
                sheet_matches.append({"row": row_idx, "cells": cells})
        if sheet_matches:
            matches[sname] = sheet_matches

        slots = fullbp_slots if sname == "FullBP" else build_slot_view(ws, fullbp_index=fullbp_index)
        matching_slots = []
        for slot in slots:
            searchable_values = build_slot_search_text(slot)
            if any(search_text in value.lower() for value in searchable_values if value):
                matching_slots.append(slot)
        if matching_slots:
            slot_matches[sname] = matching_slots

    if not matches and not slot_matches:
        print(json.dumps({
            "query": args.query,
            "matches": {},
            "slot_matches": {},
            "message": "No rows found matching the query."
        }))
    else:
        print(json.dumps({
            "query": args.query,
            "matches": matches,
            "slot_matches": slot_matches
        }, default=str, indent=2))


def cmd_update_cell(args):
    """Update a single cell identified by sheet name, row number, and column."""
    excel_path = get_excel_path()
    col_idx = resolve_col(args.col)
    row_idx = args.row
    value = coerce_value(args.value)

    bak_path = backup_file(excel_path)

    wb = load_workbook_writable(excel_path)
    ws = get_sheet(wb, args.sheet)

    if row_idx < 1 or row_idx > ws.max_row:
        print(json.dumps({"error": f"Row {row_idx} is out of range (1 to {ws.max_row})."}))
        sys.exit(1)

    col_letter = get_column_letter(col_idx)
    old_value = ws.cell(row=row_idx, column=col_idx).value
    ws.cell(row=row_idx, column=col_idx).value = value
    wb.save(str(excel_path))

    print(json.dumps({
        "status": "updated",
        "sheet": args.sheet,
        "row": row_idx,
        "col": col_letter,
        "old_value": old_value,
        "new_value": value,
        "backup": str(bak_path)
    }, default=str, indent=2))


def cmd_update_match(args):
    """Find the first row where match_col equals match_val, then update update_col."""
    excel_path = get_excel_path()
    match_col_idx = resolve_col(args.match_col)
    update_col_idx = resolve_col(args.update_col)
    update_value = coerce_value(args.update_val)

    bak_path = backup_file(excel_path)

    wb = load_workbook_writable(excel_path)
    ws = get_sheet(wb, args.sheet)

    matched_row = None
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        cell_val = row[match_col_idx - 1].value
        if cell_val is not None and str(cell_val).strip() == args.match_val.strip():
            matched_row = row_idx
            break

    if matched_row is None:
        print(json.dumps({
            "error": (
                f"No row found where column '{args.match_col}' = '{args.match_val}' "
                f"in sheet '{args.sheet}'. "
                "Use --query to check the exact value (--match-val is case-sensitive)."
            )
        }))
        sys.exit(1)

    match_col_letter = get_column_letter(match_col_idx)
    update_col_letter = get_column_letter(update_col_idx)
    old_value = ws.cell(row=matched_row, column=update_col_idx).value
    ws.cell(row=matched_row, column=update_col_idx).value = update_value
    wb.save(str(excel_path))

    print(json.dumps({
        "status": "updated",
        "sheet": args.sheet,
        "matched_row": matched_row,
        "match_col": match_col_letter,
        "match_val": args.match_val,
        "update_col": update_col_letter,
        "old_value": old_value,
        "new_value": update_value,
        "backup": str(bak_path)
    }, default=str, indent=2))


def apply_insert_slot_plan(args, plan):
    """Execute an insert-slot plan against the workbook."""
    if plan.get("blocking_issues"):
        print(json.dumps(plan, default=str, indent=2))
        sys.exit(1)

    excel_path = get_excel_path()
    bak_path = backup_file(excel_path)
    wb = load_workbook_writable(excel_path)
    context = build_workbook_context(wb)

    simpler_ws = context["simpler_ws"]
    fullbp_ws = context["fullbp_ws"]

    simpler_insert_col = plan["targets"]["simpler_insert_col"]
    fullbp_insert_col = plan["targets"]["fullbp_insert_col"]

    simpler_template_col = plan.get("style_templates", {}).get("simpler", {}).get("slot_col_index")
    fullbp_style_donor_col = plan.get("style_templates", {}).get("fullbp", {}).get("slot_col_index")
    if simpler_template_col is None:
        simpler_template_col = plan.get("templates", {}).get("simpler", {}).get("slot_col_index")
    if simpler_template_col is None:
        simpler_template_col = plan["anchor"]["simpler"]["slot_col_index"]
    if fullbp_style_donor_col is None:
        fullbp_style_donor_col = plan["templates"]["fullbp"]["slot_col_index"]

    simpler_dimensions = snapshot_column_dimensions(simpler_ws)
    fullbp_dimensions = snapshot_column_dimensions(fullbp_ws)

    simpler_ws.insert_cols(simpler_insert_col)
    fullbp_ws.insert_cols(fullbp_insert_col)

    restore_shifted_column_dimensions(simpler_ws, simpler_dimensions, simpler_insert_col)
    restore_shifted_column_dimensions(fullbp_ws, fullbp_dimensions, fullbp_insert_col)

    if simpler_template_col >= simpler_insert_col:
        simpler_template_col += 1
    if fullbp_style_donor_col >= fullbp_insert_col:
        fullbp_style_donor_col += 1

    copy_column_style(simpler_ws, simpler_template_col, simpler_insert_col)
    copy_column_style(fullbp_ws, fullbp_style_donor_col, fullbp_insert_col)

    simpler_rows = build_row_lookup(simpler_ws)
    fullbp_rows = build_row_lookup(fullbp_ws)

    for row_entry in plan.get("fullbp_row_plan", []):
        row_label = row_entry["row_label"]
        if row_label not in fullbp_rows:
            continue
        target_cell = fullbp_ws.cell(row=fullbp_rows[row_label], column=fullbp_insert_col)
        if row_entry.get("status") in PLAN_VALUE_STATUSES:
            target_cell.value = row_entry.get("value")
        elif row_entry.get("status") == "blank":
            target_cell.value = None

    dielet_flows = parse_dielet_flows(args.dielet_flow)
    simpler_ws.cell(
        row=simpler_rows[SEMANTIC_ROW_LABELS["all_flow"]],
        column=simpler_insert_col,
    ).value = args.new_flow
    for dielet in DIELET_KEYS:
        field = DIELET_FIELD_MAP[dielet]
        simpler_row = simpler_rows.get(SEMANTIC_ROW_LABELS[field])
        flow_name = dielet_flows.get(dielet)
        if simpler_row:
            simpler_ws.cell(row=simpler_row, column=simpler_insert_col).value = flow_name

    wb.save(str(excel_path))

    result = dict(plan)
    result["status"] = "inserted"
    result["backup"] = str(bak_path)
    print(json.dumps(result, default=str, indent=2))


def cmd_report_ambiguities(args):
    """Dry-run an insert-slot request and report unresolved inferred fields."""
    wb = load_workbook_readonly(get_excel_path())
    plan = plan_insert_slot(args, wb)
    if plan.get("blocking_issues"):
        plan["status"] = "blocked"
    else:
        plan["status"] = "ambiguous" if plan.get("ambiguities") else plan.get("status", "ready")
    print(json.dumps(plan, default=str, indent=2))


def cmd_insert_slot(args):
    """Insert a slot using SimplerBPView as the request anchor."""
    wb = load_workbook_readonly(get_excel_path())
    plan = plan_insert_slot(args, wb)
    required_owner = plan.get("confirmation_required")
    if required_owner and args.confirm_owner != required_owner:
        plan.setdefault("blocking_issues", []).append(
            f"Insert requires user-confirmed owner classification. Re-run with --confirm-owner {required_owner}."
        )
        plan["status"] = "blocked"
    apply_insert_slot_plan(args, plan)


def apply_delete_slot_plan(plan):
    """Execute a delete-slot plan against the workbook."""
    if plan.get("blocking_issues"):
        print(json.dumps(plan, default=str, indent=2))
        sys.exit(1)

    excel_path = get_excel_path()
    bak_path = backup_file(excel_path)
    wb = load_workbook_writable(excel_path)
    simpler_ws = get_sheet(wb, "SimplerBPView")
    fullbp_ws = get_sheet(wb, "FullBP")

    simpler_ws.delete_cols(plan["targets"]["simpler"]["slot_col_index"])
    fullbp_ws.delete_cols(plan["targets"]["fullbp"]["slot_col_index"])

    wb.save(str(excel_path))

    result = dict(plan)
    result["status"] = "deleted"
    result["backup"] = str(bak_path)
    print(json.dumps(result, default=str, indent=2))


def cmd_delete_slot(args):
    """Delete a slot using SimplerBPView as the request anchor."""
    wb = load_workbook_readonly(get_excel_path())
    plan = plan_delete_slot(args, wb)
    if args.dry_run:
        plan["status"] = "ready" if not plan.get("blocking_issues") else "blocked"
        print(json.dumps(plan, default=str, indent=2))
        return
    apply_delete_slot_plan(plan)


def cmd_programflows_sync(args, apply_changes):
    """Plan or apply ProgramFlows source updates from blueprint topology."""
    wb = load_workbook_readonly(get_excel_path())
    try:
        context = build_workbook_context(wb)
    finally:
        wb.close()

    plan = build_sync_plan(REPO_ROOT, context, args.bom, args.repo_scope)
    if apply_changes:
        print(serialize_plan(apply_sync_plan(plan)))
        return
    print(serialize_plan(plan))


# ---------------------------------------------------------------------------
# Argument parsing
# ---------------------------------------------------------------------------

def build_parser():
    parser = argparse.ArgumentParser(
        description="NVL Blueprint Excel reader and updater.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    # Read supported blueprint sheets (all cells, with colors + slot view)
  python blueprint.py --read

    # Read SimplerBPView only
    python blueprint.py --read --sheet "SimplerBPView"

    # Search for a slot anchor or dielet flow across both supported sheets
    python blueprint.py --query "STARTPCDPATMODSPKG"

    # Search within SimplerBPView only
    python blueprint.py --query "VStartPCDPost" --sheet "SimplerBPView"

    # Update a specific cell by coordinates
    python blueprint.py --update-cell --sheet "SimplerBPView" --row 15 --col E --value "STARTPCDNOM"

    # Find a row by identifier and update a different column
    python blueprint.py --update-match --sheet "FullBP" --match-col A --match-val "FLOWDEF:ALL:FlowType" --update-col J --update-val "SubParFlow"

    # Dry-run an insert to see inferred FullBP metadata and ambiguities
    python blueprint.py --report-ambiguities --anchor-flow "STARTPCDPATMODSPKG" --position after --new-flow "StartPost" --mode parallel --dielet-flow CPU=VStartCPUPost --dielet-flow PCD=VStartPCDPost --explicit-dielet-subflows

    # Insert a new parallel slot after STARTPCDPATMODSPKG
    python blueprint.py --insert-slot --anchor-flow "STARTPCDPATMODSPKG" --position after --new-flow "StartPost" --mode parallel --dielet-flow CPU=VStartCPUPost --dielet-flow PCD=VStartPCDPost --explicit-dielet-subflows

    # Dry-run deletion of an existing slot
    python blueprint.py --delete-slot --target-flow "STARTPCDPATMODSPKG" --dry-run
        """
    )

    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument(
        "--read", action="store_true",
        help="Read and dump the Excel file as JSON (values + fill colors)."
    )
    mode.add_argument(
        "--query", metavar="TEXT",
        help="Search for rows where any cell contains TEXT (case-insensitive)."
    )
    mode.add_argument(
        "--update-cell", action="store_true",
        help="Update a single cell by row number and column. Requires --sheet, --row, --col, --value."
    )
    mode.add_argument(
        "--update-match", action="store_true",
        help="Find a row by matching a column value and update another column. "
             "Requires --sheet, --match-col, --match-val, --update-col, --update-val."
    )
    mode.add_argument(
        "--insert-slot", action="store_true",
        help="Insert a new slot anchored on a SimplerBPView FLOWDEF:ALL:Flow value."
    )
    mode.add_argument(
        "--report-ambiguities", action="store_true",
        help="Dry-run an insert-slot request and report inferred values plus ambiguities."
    )
    mode.add_argument(
        "--delete-slot", action="store_true",
        help="Delete a slot anchored on a SimplerBPView FLOWDEF:ALL:Flow value."
    )
    mode.add_argument(
        "--plan-programflows-sync", action="store_true",
        help="Dry-run blueprint-driven updates for ProgramFlows Python sources."
    )
    mode.add_argument(
        "--apply-programflows-sync", action="store_true",
        help="Apply blueprint-driven updates for ProgramFlows Python sources."
    )

    parser.add_argument(
        "--sheet", metavar="NAME", default=None,
        help="Sheet name to operate on. For --read/--query: defaults to all sheets. "
             "For --update-*: required."
    )

    # --update-cell arguments
    cell_group = parser.add_argument_group("--update-cell arguments")
    cell_group.add_argument("--row", metavar="N", type=int, help="Row number (1-based).")
    cell_group.add_argument("--col", metavar="COL", help="Column letter (A, B, ...) or 1-based integer.")
    cell_group.add_argument("--value", metavar="VAL", help="New cell value (auto-detected as int, float, or string).")

    # --update-match arguments
    match_group = parser.add_argument_group("--update-match arguments")
    match_group.add_argument("--match-col", metavar="COL", dest="match_col", help="Column to search for the identifier.")
    match_group.add_argument("--match-val", metavar="VAL", dest="match_val", help="Exact value to match (case-sensitive).")
    match_group.add_argument("--update-col", metavar="COL", dest="update_col", help="Column to update in the matched row.")
    match_group.add_argument("--update-val", metavar="VAL", dest="update_val", help="New value to write in the matched row.")

    slot_group = parser.add_argument_group("high-level slot arguments")
    slot_group.add_argument("--anchor-flow", metavar="FLOW", help="Existing SimplerBPView FLOWDEF:ALL:Flow anchor.")
    slot_group.add_argument("--position", choices=["before", "after"], help="Insert position relative to the anchor flow.")
    slot_group.add_argument("--new-flow", metavar="FLOW", help="New FLOWDEF:ALL:Flow value to insert.")
    slot_group.add_argument("--mode", choices=["serial", "parallel"], help="Whether the new slot is serial or parallel.")
    slot_group.add_argument(
        "--dielet-flow", action="append", default=[], metavar="DIELET=FLOW",
        help="Repeatable dielet flow assignment, for example CPU=VStartCPUPost."
    )
    slot_group.add_argument(
        "--explicit-dielet-subflows", action="store_true",
        help="Required when populating FLOWDEF:<IP>:Flow child rows via --dielet-flow."
    )
    slot_group.add_argument("--scope", metavar="VALUE", help="Optional explicit FLOWDEF:ALL:Scope override.")
    slot_group.add_argument("--port-range", dest="port_range", metavar="VALUE", help="Optional explicit FLOWDEF:ALL:PortRange override.")
    slot_group.add_argument("--dielet-type", dest="dielet_type", metavar="VALUE", help="Optional explicit FLOWDEF:ALL:DieletType override.")
    slot_group.add_argument(
        "--confirm-owner", choices=["shared", "dielet"],
        help="Required for --insert-slot after user confirmation of inferred ownership classification."
    )
    slot_group.add_argument("--target-flow", metavar="FLOW", help="Existing SimplerBPView FLOWDEF:ALL:Flow value to delete.")
    slot_group.add_argument("--occurrence", metavar="N", type=int, help="1-based occurrence to use when a flow name appears more than once.")
    slot_group.add_argument("--dry-run", action="store_true", help="Plan the operation without writing workbook changes.")

    programflows_group = parser.add_argument_group("programflows sync arguments")
    programflows_group.add_argument(
        "--bom", action="append", default=[], metavar="BOM",
        help="Repeatable BOM name under POR_TP, for example Class_NVL_S28C."
    )
    programflows_group.add_argument(
        "--repo-scope", choices=["dielet", "shared", "both"], default="both",
        help="Which ProgramFlows source set to update. Defaults to both."
    )

    return parser


def main():
    parser = build_parser()
    args = parser.parse_args()

    if args.read:
        cmd_read(args)

    elif args.query is not None:
        cmd_query(args)

    elif args.update_cell:
        missing = [
            f"--{f}" for f, v in [
                ("sheet", args.sheet), ("row", args.row),
                ("col", args.col), ("value", args.value)
            ] if v is None
        ]
        if missing:
            parser.error(f"--update-cell requires: {', '.join(missing)}")
        cmd_update_cell(args)

    elif args.update_match:
        missing = [
            f"--{f}" for f, v in [
                ("sheet", args.sheet), ("match-col", args.match_col),
                ("match-val", args.match_val), ("update-col", args.update_col),
                ("update-val", args.update_val)
            ] if v is None
        ]
        if missing:
            parser.error(f"--update-match requires: {', '.join(missing)}")
        cmd_update_match(args)

    elif args.delete_slot:
        if args.target_flow is None:
            parser.error("--delete-slot requires: --target-flow")
        cmd_delete_slot(args)

    elif args.plan_programflows_sync or args.apply_programflows_sync:
        if not args.bom:
            parser.error("ProgramFlows sync requires at least one --bom entry.")
        cmd_programflows_sync(args, apply_changes=args.apply_programflows_sync)

    elif args.report_ambiguities or args.insert_slot:
        missing = [
            f"--{f}" for f, v in [
                ("anchor-flow", args.anchor_flow),
                ("position", args.position),
                ("new-flow", args.new_flow),
                ("mode", args.mode),
            ] if v is None
        ]
        if missing:
            parser.error(f"High-level slot operations require: {', '.join(missing)}")
        if args.mode == "parallel" and not args.dielet_flow:
            parser.error("Parallel slot insertion requires at least one --dielet-flow entry.")
        if args.dielet_flow and not args.explicit_dielet_subflows:
            parser.error(
                "Dielet-specific child flows require --explicit-dielet-subflows. "
                "Leave FLOWDEF:<IP>:Flow rows empty unless the user explicitly requested those child subflows."
            )
        if args.report_ambiguities:
            cmd_report_ambiguities(args)
        else:
            cmd_insert_slot(args)


if __name__ == "__main__":
    main()
