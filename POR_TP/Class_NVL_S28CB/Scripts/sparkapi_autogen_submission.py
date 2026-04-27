import os
import sys
import subprocess
import json
import pandas as pd
import shutil
import re
from datetime import datetime
from openpyxl import load_workbook
import logging
import time
import win32com.client


def setup_logging(base_directory):
    # Configure logging to use the base directory for the log file
    log_file_path = os.path.join(base_directory, 'submission_log.txt')
    logging.basicConfig(
        filename=log_file_path,  # Log file name
        level=logging.DEBUG,  # Log level
        format='%(asctime)s - %(levelname)s - %(message)s'  # Log format
    )
    logging.info(f"Logging initialized. Log file path: {log_file_path}")


class ProcessMultiTPExcel:
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path

    def update_excel_multi(self, directory_path, tp_rev):
        # Print the user input path
        print(f"Processing Directory_Path: {directory_path}")

        # Load the workbook and select the sheet
        workbook = load_workbook(self.excel_file_path)
        sheet = workbook['VPO_Jobs']

        # Iterate over the rows and update the TPRevision column
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            if row[1].value == directory_path:  # Check if the DirectoryPath matches
                additional_tags = row[4].value  # AdditionalTags is now the fifth column (index 4)
                if additional_tags and additional_tags.strip():  # Check if AdditionalTags is not blank
                    row[2].value = f"{tp_rev}_{additional_tags.strip()}"  # Update TPRevision
                else:
                    row[2].value = tp_rev  # Update TPRevision without AdditionalTags

        # Save the workbook
        workbook.save(self.excel_file_path)

    def process_excel_multi(self):
        # Read the Excel file to get the 'VPO_Jobs' sheet
        df_jobs = pd.read_excel(self.excel_file_path, sheet_name='VPO_Jobs')

        # Iterate over each row to process the directory path and determine TP_REV
        for index, row in df_jobs.iterrows():
            directory_path = row['DirectoryPath']
            
            # Process user input to determine TP_REV
            user_input_processor = ProcessUserInput(directory_path)
            try:
                tp_rev = user_input_processor.determine_tp_rev()
                # Update the Excel file with the determined TP_REV
                self.update_excel_multi(directory_path, tp_rev)
                print(f"Updated TP_REV for directory {directory_path}: {tp_rev}")
            except ValueError as e:
                print(f"Error determining TP_REV for directory {directory_path}: {e}")
                continue


class ProcessUserInput:
    def __init__(self, directory_path):
        self.directory_path = directory_path

    def determine_tp_rev(self):
        # Check if the directory exists
        if os.path.isdir(self.directory_path):
            print(f"Directory exists: {self.directory_path}")
        else:
            raise ValueError(f"Directory does not exist: {self.directory_path}")

        # Define possible paths to the usrv files
        paths = [
            os.path.join(self.directory_path, 'Shared', 'BaseInputs', 'UservarDefinitions.usrv'),
            os.path.join(self.directory_path, 'Shared', 'BaseInputs', 'Common', 'Common_Files', 'common.usrv')
        ]

        # Search for TP_REV in the files
        for path in paths:
            if os.path.exists(path):
                print(f"File exists: {path}")  # Print the file path if it exists
                with open(path, 'r') as file:
                    for line in file:
                        if 'Const String TPName' in line:
                            # Extract TP_REV from the line
                            tp_rev = line.split('"')[1]
                            print(f"Full TP_REV before trim: {tp_rev}")  # Print the full TP_REV before trimming
                            # Extract the desired substring from TP_REV
                            trimmed_tp_rev = tp_rev[10:15]  # Adjust indices based on your requirement
                            print(f"Trimmed TP_REV: {trimmed_tp_rev}")  # Print the trimmed TP_REV
                            return trimmed_tp_rev
            else:
                print(f"File does not exist: {path}")  # Print a warning if the file does not exist

        raise ValueError("TP_REV not found in the specified files.")

    def update_excel(self, excel_file_path, directory_path, tp_rev):
        # Print the user input path
        print(f"User input Directory_Path: {directory_path}")

        # Load the workbook and select the sheet
        workbook = load_workbook(excel_file_path)
        sheet = workbook['VPO_Jobs']

        # Iterate over the rows and update the TPRevision column
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            additional_tags = row[4].value  # AdditionalTags is now the fifth column (index 4)
            if additional_tags and additional_tags.strip():  # Check if AdditionalTags is not blank
                row[2].value = f"{tp_rev}_{additional_tags.strip()}"  # Update TPRevision
            else:
                row[2].value = tp_rev  # Update TPRevision without AdditionalTags

        # Update the DirectoryPath column
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            row[1].value = directory_path  # Assuming 'DirectoryPath' is the second column

        # Save the workbook
        workbook.save(excel_file_path)


class SubmissionRead:
    def __init__(self, excel_file_path, current_path):
        self.excel_file_path = self.convert_path(excel_file_path)
        self.current_path = current_path

    def convert_path(self, path):
        # Convert backslashes to forward slashes for cross-platform compatibility
        return path.replace('\\', '/')

    def read_excel_file(self, file_path, sheet_name):
        # Read the specified sheet from the Excel file
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        return df

    def find_files(self, base_path, tp_type):
        # Generic regular expressions to match the desired file patterns
        stpl_pattern = re.compile(r'.*SubTestPlan.*\.stpl$')
        plist_pattern = re.compile(r'.*PLIST_ALL_CLASS_.*\.plist\.xml$')

        stpl_path = None
        plist_path = None

        search_path = os.path.join(base_path, tp_type)

        for root, dirs, files in os.walk(search_path):
            for file in files:
                full_path = os.path.join(root, file)
                full_path = self.convert_path(full_path)  # Ensure path uses forward slashes
                if stpl_pattern.match(full_path):
                    stpl_path = full_path.replace(base_path + '/', '')
                elif plist_pattern.match(full_path):
                    plist_path = full_path.replace(base_path + '/', '')

        if not stpl_path or not plist_path:
            raise ValueError("Could not find STPL or PLIST files with the specified patterns.")

        return stpl_path, plist_path

    def determine_tp_type(self, directory_path):
        # Automatically determine TP_TYPE by checking the directory structure
        for root, dirs, files in os.walk(directory_path):
            if 'POR_TP' in dirs:
                return 'POR_TP'
            elif 'ENG_TP' in dirs:
                return 'ENG_TP'
        raise ValueError("Could not determine TP_TYPE. Please ensure the directory contains either 'POR_TP' or 'ENG_TP'.")

    def extract_excel_inputs(self, df_jobs, df_static, df_submission_option, df_site_labs, index):
        # Extract values from the VPO_Jobs sheet
        Directory_Path = df_jobs.at[index, 'DirectoryPath']
        TP_REV = df_jobs.at[index, 'TPRevision']
        
        # Extract Ref_VPO and handle NaN values properly
        Ref_VPO = df_jobs.at[index, 'Ref_VPO']
        if pd.isna(Ref_VPO) or Ref_VPO == '' or str(Ref_VPO).strip() == '' or str(Ref_VPO).lower() == 'nan':
            Ref_VPO = "N/A"  # Set default value instead of nan
        else:
            Ref_VPO = str(Ref_VPO).strip()  # Convert to string and strip whitespace
        
        AdditionalTags = df_jobs.at[index, 'AddtionalTags']  # Note: keeping the typo as it appears in your column header
        OverwritingPartType = df_jobs.at[index, 'OverwritingPartType']
        FusedPart = df_jobs.at[index, 'FusedPart']
        QDFS = df_jobs.at[index, 'QDFS']
        hri = df_jobs.at[index, 'hri']
        mrv = df_jobs.at[index, 'mrv']
        recipe_config = df_jobs.at[index, 'recipe_config']
        SourceLot = df_jobs.at[index, 'SourceLot']
        VisualIDPath = df_jobs.at[index, 'VisualIDPath']
        NumberofUnitsToRun = df_jobs.at[index, 'NumberofUnitsToRun']
        Stepping = df_jobs.at[index, 'Stepping']
        PartType = df_jobs.at[index, 'PartType']
        Location = df_jobs.at[index, 'Location']
        EngID = df_jobs.at[index, 'EngID']
        AdditionalComments = df_jobs.at[index, 'AdditionalComments']  # Extract AdditionalComments
        SubmissionOption = df_jobs.at[index, 'SubmissionOption']

        # Extract values from the StaticInputs sheet (always first row)
        Site_Chosen = df_static.at[0, 'SiteChosen']
        SSID = df_static.at[0, 'SSID']
        Contact = df_static.at[0, 'Contact']

        # Extract TAG and SpecialInstruction from the SubmissionOption sheet
        submission_info = df_submission_option.loc[df_submission_option['SubmissionOption'] == SubmissionOption]
        TAG = submission_info['TAG'].values[0]
        SpecialInstruction = submission_info['SpecialInstruction'].values[0]

        # Extract segment, team, and lab information from the Site-Labs sheet
        site_info = df_site_labs.loc[df_site_labs['Selector'] == Site_Chosen]
        Segment = site_info['Segment'].values[0]
        Team = site_info['Team'].values[0]
        Labs = site_info['Labs'].values[0]

        # Strip whitespace from recipe_config
        recipe_config = str(recipe_config).strip()

        # Validate and process inputs
        if not os.path.isdir(Directory_Path):
            raise ValueError("Path is Invalid!")
        BaseTP = self.convert_path(Directory_Path)

        # Automatically determine TP_TYPE
        TP_TYPE = self.determine_tp_type(Directory_Path)

        if OverwritingPartType == "Y":
            OveridePartType = "True"
        elif OverwritingPartType == "N":
            OveridePartType = "False"
        else:
            raise ValueError("Invalid OverwritingPartType value. Please use 'Y' or 'N'.")

        if FusedPart == "N":
            QDFS = "XXXX"

        if recipe_config == "1":
            recipe = "1"
        elif recipe_config == "2":
            recipe = "2"
        else:
            raise ValueError("Invalid recipe_config value. Please use '1' or '2'.")

        # Set default values for hri and mrv if specified
        if hri == "Default":
            hri = "DEFAULT_HRI"
        if mrv == "Default":
            mrv = "000000000000"

        return (BaseTP, Segment, Team, TP_TYPE, TP_REV, OveridePartType, QDFS, hri, mrv, recipe,
                SourceLot, VisualIDPath, NumberofUnitsToRun, Stepping, PartType, Location, EngID, TAG, SpecialInstruction, Contact, SSID, SubmissionOption, Labs, AdditionalComments, Ref_VPO)

    def process_excel(self):
        # Process the Excel file and create a temporary file with processed data
        if not os.path.isfile(self.excel_file_path):
            print("Invalid file path. Please ensure the file exists.")
            return

        df_jobs = self.read_excel_file(self.excel_file_path, sheet_name='VPO_Jobs')
        df_static = self.read_excel_file(self.excel_file_path, sheet_name='StaticInputs')
        df_site_labs = self.read_excel_file(self.excel_file_path, sheet_name='Site-Labs')
        df_submission_option = self.read_excel_file(self.excel_file_path, sheet_name='SubmissionOption')

        # Get SSID from Excel file
        ssid = df_static.at[0, 'SSID']

        # Use single vpo_run directory
        vpo_run_directory = os.path.join(self.current_path, 'Submission_History', ssid, 'vpo_run')
        os.makedirs(vpo_run_directory, exist_ok=True)

        # Set up logging
        setup_logging(vpo_run_directory)

        # Open the temporary file for writing
        temp_file_path = os.path.join(vpo_run_directory, 'vpo_jobs.temp')
        logging.info(f"Creating temporary file at: {temp_file_path}")

        with open(temp_file_path, 'w') as temp_file:
            # Iterate over each row in the VPO_Jobs DataFrame
            for index in range(len(df_jobs)):
                try:
                    inputs = self.extract_excel_inputs(df_jobs, df_static, df_submission_option, df_site_labs, index)
                    
                    # Find STPL and PLIST paths using regular expressions
                    STPL, PLIST = self.find_files(inputs[0], inputs[3])  # BaseTP and TP_TYPE

                    # Extract grp_displayname_tprev from TP_REV
                    grp_displayname_tprev = inputs[4][:5]  # TP_REV is the fifth element in inputs

                    # Log extracted values
                    logging.info(f"Processed values for row {index}: {inputs}")
                    # Write to temp file
                    temp_file.write(f"Processed values for row {index}:\n")
                    for input_name, input_value in zip(["BaseTP", "Segment", "Team", "TP_TYPE", "STPL", "PLIST", "TP_REV", "OveridePartType", "QDFS", "hri", "mrv", "recipe",
                                                        "SourceLot", "VisualIDPath", "NumberofUnitsToRun", "Stepping", "PartType", "Location", "EngID", "TAG", "SpecialInstruction", "Contact", "SSID", "SubmissionOption", "Labs", "AdditionalComments", "Ref_VPO", "grp_displayname_tprev"], 
                                                       [*inputs[:4], STPL, PLIST, *inputs[4:], grp_displayname_tprev]):

                        #################################################
                        # logging.debug(f"{input_name}: {input_value}") # Debug Printing Enable if needed
                        #################################################
                        temp_file.write(f"{input_name}: {input_value}\n")
                    temp_file.write("\n")  # Add a newline between entries

                except ValueError as e:
                    logging.error(f"Error processing row {index}: {e}")

        # Copy the Excel file to the vpo_run directory
        shutil.copy(self.excel_file_path, vpo_run_directory)

        # Return the path to the temporary file, the base directory, and the SSID
        return temp_file_path, vpo_run_directory, ssid


class GenerateVpoJobsReport:
    def __init__(self, log_file_path, vpo_run_directory):
        self.log_file_path = log_file_path
        self.vpo_run_directory = vpo_run_directory

    def read_log_file(self):
        # Log the log file path
        logging.info(f"Reading log file from: {self.log_file_path}")

        # Read the log file and parse the data into a list of dictionaries
        data_entries = []
        contact_emails_set = set()  # Initialize a set to collect unique emails
        with open(self.log_file_path, 'r') as file:
            entry = {}
            for line in file:
                line = line.strip()
                # Check if the line contains a data entry
                if "INFO - {" in line:
                    # Extract the JSON-like string from the log entry
                    json_str = line.split("INFO - ", 1)[1]
                    try:
                        # Convert the JSON-like string to a dictionary
                        entry = eval(json_str)
                        data_entries.append(entry)
                        # Extract emails from the Contact field
                        contact_emails = entry.get('Contact', '')
                        contact_emails_set.update(contact_emails.split(','))
                    except Exception as e:
                        logging.error(f"Error parsing log entry: {e}")
                elif "Processed values for row" in line:
                    if entry:
                        data_entries.append(entry)
                    entry = {}

        # Log the extracted data entries
        logging.info("Extracted data entries:")
        for entry in data_entries:
            logging.info(entry)

        return data_entries, contact_emails_set

    def generate_report_info(self):
        # Read data from the log file
        data_entries, contact_emails_set = self.read_log_file()

        # Check if data_entries is empty or missing expected keys
        if not data_entries or 'BaseTP' not in data_entries[0]:
            logging.error("Missing 'BaseTP' in data entries. Please check the log file format.")
            raise KeyError("Missing 'BaseTP' in data entries. Please check the log file format.")

        # Extract TP name from BaseTP
        base_tp = data_entries[0]['BaseTP']
        logging.info(f"BaseTP value is {base_tp}")

        # Extract TP name using string manipulation
        tp_name = base_tp.split('/')[-1]  # Get the last segment after the last '/'

        # Generate report
        report_path = os.path.join(self.vpo_run_directory, 'report_info.txt')

        with open(report_path, 'w') as report_file:
            # Write TestProgram and ContactEmail to both files
            report_file.write(f'TestProgram: {tp_name}\n')
            report_file.write(f'ContactEmail: {",".join(contact_emails_set)}\n')

            # Initialize a counter dictionary for each SubmissionOption
            counters = {}

            # Sort entries by SubmissionOption to ensure correct ordering
            sorted_entries = sorted(data_entries, key=lambda x: x['SubmissionOption'])

            for entry in sorted_entries:
                submission_option = entry['SubmissionOption']
                if submission_option not in counters:
                    counters[submission_option] = 0  # Initialize counter for each SubmissionOption

                counters[submission_option] += 1  # Increment counter for the current SubmissionOption

                # Extract tag_prefix from PLIST path
                plist_segments = entry['PLIST'].split('/')
                if len(plist_segments) > 2:
                    tag_prefix = plist_segments[2].split('PLIST_ALL_')[1].split('.')[0]
                else:
                    raise ValueError("PLIST path does not have the expected format.")

                display_name = f"{submission_option} [{counters[submission_option]:02}] {entry['TP_REV']} {entry['Location']} {entry['EngID']}"  # Construct DisplayName
                ref_vpo = entry['Ref_VPO']  # Get Ref_VPO from the entry
                source_lot = entry['SourceLot']
                num_units = entry['NumberofUnitsToRun']
                location = entry['Location']
                eng_id = entry['EngID']
                tag = f"{tag_prefix}_{entry['TAG']}_{entry['TP_REV']}"
                
                # Write in the format expected by GetData.ps1: DisplayName-Lot-Qty-Operation-EngID-Tags-Ref_VPO
                report_file.write(f'{display_name}-{source_lot}-{num_units}-{location}-{eng_id}-{tag}-{ref_vpo}\n')
                logging.info(f"Written to report: DisplayName={display_name}, Ref_VPO={ref_vpo}")

        logging.info(f"Report generated at: {report_path}")


class SparkGenApi:
    def __init__(self, temp_file_path, vpo_run_directory, tp_site, common_path):
        self.temp_file_path = temp_file_path
        self.vpo_run_directory = vpo_run_directory
        self.tp_site = tp_site  # Store tp_site as an instance variable
        self.common_path = common_path  # Store common_path as an instance variable

    def read_temp_file(self):
        # Check if the temporary file exists before reading
        if not os.path.exists(self.temp_file_path):
            logging.error(f"Temporary file not found: {self.temp_file_path}")
            raise FileNotFoundError(f"Temporary file not found: {self.temp_file_path}")

        # Read the temporary file and parse the data into a list of dictionaries
        data = []
        with open(self.temp_file_path, 'r') as file:
            entry = {}
            for line in file:
                line = line.strip()
                if line.startswith("Processed values for row"):
                    if entry:
                        data.append(entry)
                    entry = {}
                elif line:
                    key, value = line.split(": ", 1)
                    entry[key] = value
            if entry:
                data.append(entry)

        logging.info("Data entries read from temp file:")
        for entry in data:
            logging.info(entry)

        return data

    def convert_base_tp_to_network_path(self, base_tp):
        # Convert the base TP path to a network path based on the tp_site argument
        if 'I:/' in base_tp:
            if self.tp_site.upper() == 'PG':
                base_tp = base_tp.replace('I:/', '\\\\gar.corp.intel.com\\ec\\proj\\mdl\\pg\\intel\\')
            elif self.tp_site.upper() == 'JF':
                base_tp = base_tp.replace('I:/', '\\\\amr.corp.intel.com\\ec\\proj\\mdl\\jf\\intel\\')
            elif self.tp_site.upper() == 'FM':
                base_tp = base_tp.replace('I:/', '\\\\amr.corp.intel.com\\ec\\proj\\mdl\\fm\\intel\\')
            else:
                raise ValueError("Invalid tp_site value. Please use 'PG' or 'JF' pr 'FM'.")
            base_tp = base_tp.replace('/', '\\')
        return base_tp

    def read_visual_ids(self, visual_id_path):
        # Read the Visual IDs from the specified file path
        with open(visual_id_path, 'r') as file:
            data = [line.strip() for line in file.readlines()]
        return data

    def remove_none_values(self, d):
        """Recursively remove None values from dictionaries."""
        if isinstance(d, dict):
            return {k: self.remove_none_values(v) for k, v in d.items() if v is not None}
        elif isinstance(d, list):
            return [self.remove_none_values(v) for v in d if v is not None]
        else:
            return d

    def create_submission_json(self, data_entries, submission_type):
        # Create a JSON file for the submission based on the data entries and submission type
        submission_file = os.path.join(self.vpo_run_directory, f'submission_vpo_{submission_type}.json')

        # Remove old submission file if exists
        if os.path.exists(submission_file):
            os.remove(submission_file)

        experiments = []
        counter = 0  # Initialize counter for each submission type
        for data_entry in data_entries:
            counter += 1
            # Convert BaseTP to network path
            base_tp_network_path = self.convert_base_tp_to_network_path(data_entry['BaseTP'])

            # Prepare the stplDirectory
            path_stpl = data_entry['STPL'].split('/')[0] + '\\' + data_entry['STPL'].split('/')[1]
            stpl_directory = base_tp_network_path + '\\' + path_stpl

            # Check if VisualIDPath is valid
            use_vid = False
            visual_ids = []
            visual_id_path = data_entry.get('VisualIDPath', None)
            if visual_id_path and pd.notna(visual_id_path) and os.path.exists(visual_id_path):
                use_vid = True
                visual_ids = self.read_visual_ids(visual_id_path)

            # Determine recipe mode
            recipe_mode = data_entry['recipe']
            recipe_directory = os.path.join(base_tp_network_path, 'astra') if recipe_mode == "2" else None

            # Extract the desired portion from the PLIST path
            plist_segments = data_entry['PLIST'].split('/')
            if len(plist_segments) > 2:
                # Extract the part after 'PLIST_ALL_' in the third segment
                tag_prefix = plist_segments[2].split('PLIST_ALL_')[1].split('.')[0]
            else:
                raise ValueError("PLIST path does not have the expected format.")

            # Set experimentType and activityType based on submission type
            if submission_type in ["MV", "MV-PFused"]:
                experiment_type = "Engineering"
                activity_type = "ModuleValidation"
            else:
                experiment_type = "Correlation"
                activity_type = "Correlation"

            # Combine SpecialInstruction and AdditionalComments
            combined_comments = f"{data_entry['SpecialInstruction']}\n{data_entry.get('AdditionalComments', '')}"  # Use .get() to avoid KeyError

            # Prepare the experiment data
            experiment = {
                "displayName": f"{submission_type} [{counter:02}] {data_entry['TP_REV']} {data_entry['Location']} {data_entry['EngID']}",
                "experimentType": experiment_type,
                "activityType": activity_type,
                "bomGroupName": tag_prefix,
                "step": data_entry['Stepping'],
                "tplFileName": None if recipe_mode == "2" else "..\\..\\BaseTestPlan.tpl",
                "stplFileName": None if recipe_mode == "2" else data_entry['STPL'].split('/')[2],
                "environmentFileName": "EnvironmentFile.env" if submission_type == "Post-Fuse" or recipe_mode != "2" else None,
                "plistAllFileName": None if recipe_mode == "2" else data_entry['PLIST'].split('/')[2],
                "testTimeInSecPerUnit": 400,
                "retestRate": 50,
                "lab": data_entry['Labs'],
                "material": {
                    "materialIssue": {
                        "materialIssueIsRequired": False
                    },
                    "units": [{"visualId": vid} for vid in visual_ids] if use_vid else None,
                    "lots": [
                        {
                            "name": data_entry['SourceLot'],
                            "numOfUnitsToRun": int(data_entry['NumberofUnitsToRun'])
                        }
                    ] if not use_vid else None,
                    "partTypeOverride": data_entry['PartType'] if data_entry['OveridePartType'] == "True" else None,
                    "sspec": data_entry['QDFS'] if submission_type == "Post-Fuse" or submission_type == "MV-PFused" else None
                },
                "conditions": [
                    {
                        "operation": data_entry['Location'],
                        "engineeringId": data_entry['EngID'],
                        "dieSelection": "NA",
                        "comment": combined_comments,  # Use combined comments
                        "moveUnits": "Good"
                    }
                ],
                "flexbom": {
                    "hri": data_entry['hri'],
                    "mrv": data_entry['mrv']
                },
                "experimentState": "Ready",
                "tags": [f"{tag_prefix}_{data_entry['TAG']}_{data_entry['TP_REV']}"],
                "ContactEmails": [data_entry['Contact']],
                "recipe": {
                    "recipeSource": "ByoUseAsIs" if recipe_mode == "2" else "TpGenerated",
                    "recipePath": recipe_directory if recipe_mode == "2" else None
                }
            }

            experiments.append(experiment)

        grp_displayname_tprev = data_entries[0]['grp_displayname_tprev']

        # Prepare the JSON data
        json_data = {
            "segment": data_entries[0]['Segment'],
            "team": data_entries[0]['Team'],
            "displayName": f"{submission_type} {grp_displayname_tprev}",
            "stplDirectory": stpl_directory,
            "experiments": experiments
        }

        # Remove None values from the JSON data
        json_data_cleaned = self.remove_none_values(json_data)

        # Write the cleaned JSON data to the file
        with open(submission_file, 'w') as f:
            json.dump(json_data_cleaned, f, indent=4)

        logging.info(f"{submission_type} submission JSON is successfully created.")

    def submit_to_spark(self, submission_file):
        # Determine the site-specific prefix
        if self.tp_site.upper() == "PG":
            site_prefix = "\\\\gar\\ec\\proj\\mdl\\pg\\"
        elif self.tp_site.upper() in ["JF", "FM"]:
            site_prefix = "\\\\amr\\ec\\proj\\mdl\\jf\\"
        else:
            raise ValueError("Invalid tp_site value. Please use 'PG', 'JF', or 'FM'.")

        # Call the PowerShell script to submit the JSON file to SPARK
        logging.info("Calling the Spark submission file...")
        output_directory = self.vpo_run_directory

        # Always use Default path
        script_path = site_prefix + self.common_path.replace('/', '\\')

        subprocess.call([
            "powershell",
            "-ExecutionPolicy", "Bypass",
            #f"C:\\Users\\samrenji\\Desktop\\SPARK\\submission_MV.ps1",
            f"{script_path}\\submission_MV.ps1",
            "-JsonFilePath", submission_file,
            "-OutputDirectory", output_directory
        ])

    def process_submissions(self):
        # Get the path to the temporary file from command-line arguments
        data_entries = self.read_temp_file()

        # Group entries by submission type
        submission_groups = {}
        for data_entry in data_entries:
            submission_option = data_entry['SubmissionOption']
            if submission_option not in submission_groups:
                submission_groups[submission_option] = []
            submission_groups[submission_option].append(data_entry)

        # Process each group
        for submission_type, entries in submission_groups.items():
            self.create_submission_json(entries, submission_type)
            self.submit_to_spark(os.path.join(self.vpo_run_directory, f'submission_vpo_{submission_type}.json'))


class GetVpoNumFromSpark:
    def __init__(self, report_file_path, submission_file_path, vpo_run_directory, tp_site, common_path):
        self.report_file_path = report_file_path
        self.submission_file_path = submission_file_path
        self.vpo_run_directory = vpo_run_directory
        self.tp_site = tp_site  # Store tp_site as an instance attribute
        self.common_path = common_path  # Store common_path as an instance variable

    def run_get_data_script(self):
        # Determine the site-specific prefix
        if self.tp_site.upper() == "PG":
            site_prefix = "\\\\gar\\ec\\proj\\mdl\\pg\\"
        elif self.tp_site.upper() in ["JF", "FM"]:
            site_prefix = "\\\\amr\\ec\\proj\\mdl\\jf\\"
        else:
            raise ValueError("Invalid tp_site value. Please use 'PG', 'JF', or 'FM'.")

        # Call the PowerShell script to get VPO numbers from SPARK
        logging.info("Calling the GetData PowerShell script...")

        # Always use Default path
        script_path = site_prefix + self.common_path.replace('/', '\\')
        
        subprocess.call([
            "powershell",
            "-ExecutionPolicy", "Bypass",
            #f"C:\\Users\\samrenji\\Desktop\\SPARK\\GetData.ps1",
            f"{script_path}\\GetData.ps1",
            "-SubmissionFilePath", self.submission_file_path,
            "-ReportFilePath", self.report_file_path,
            "-OutputDirectory", self.vpo_run_directory
        ])


class ProcessOutput:
    def __init__(self, vpo_run_directory):
        self.vpo_run_directory = vpo_run_directory
        self.output_file_path = os.path.join(self.vpo_run_directory, 'output.txt')

    def clean_line(self, line):
        return ''.join(line.replace('\x00', '').split())

    def separate_output(self):
        """Separate output.txt into correlation_table.txt and mv_table.txt"""
        correlation_file_path = os.path.join(self.vpo_run_directory, 'correlation_table.txt')
        mv_file_path = os.path.join(self.vpo_run_directory, 'mv_table.txt')

        try:
            with open(self.output_file_path, 'r', encoding='utf8', errors='ignore') as file:
                data = file.read().splitlines()

            # Clean the data
            data = [self.clean_line(line) for line in data if line.strip()]
            logging.info(f"Read {len(data)} lines from output.txt")

        except Exception as e:
            logging.error(f"Error reading output file: {e}")
            return

        with open(correlation_file_path, 'w', encoding='utf8') as correlation_file, \
             open(mv_file_path, 'w', encoding='utf8') as mv_file:

            # Write headers to both files
            for line in data:
                if line.startswith('TestProgram:') or line.startswith('ContactEmail:'):
                    correlation_file.write(line + '\n')
                    mv_file.write(line + '\n')
                    logging.info(f"Wrote to both files: {line}")

            # Separate data by submission type
            for line in data:
                if line.startswith('DisplayName:Correlation') or line.startswith('DisplayName:Pre-Fuse') or line.startswith('DisplayName:Post-Fuse'):
                    correlation_file.write(line + '\n')
                    logging.info(f"Wrote to correlation_table.txt: {line}")
                elif line.startswith('DisplayName:MV') or line.startswith('DisplayName:MV-PFused'):
                    mv_file.write(line + '\n')
                    logging.info(f"Wrote to mv_table.txt: {line}")

        logging.info(f"Separated output into {correlation_file_path} and {mv_file_path}")

    def process_output(self):
        """Main processing method"""
        logging.info("Processing output.txt and separating into correlation and MV tables.")
        self.separate_output()
        logging.info("Output processing completed.")


class EmailVpoNumber:
    def __init__(self, vpo_run_directory):
        self.vpo_run_directory = vpo_run_directory
        self.mv_table_file = os.path.join(vpo_run_directory, 'mv_table.txt')
        self.correlation_table_file = os.path.join(vpo_run_directory, 'correlation_table.txt')

    def read_file(self, file_path):
        try:
            with open(file_path, 'r', encoding='utf8', errors='ignore') as file:
                data = file.readlines()

            data = [line.replace('\x00', '') for line in data]
            data = [line.strip() for line in data if line.strip()]

            return data
        except Exception as e:
            logging.error(f"Error reading file {file_path}: {e}")
            return []

    def extract_emails(self, data):
        emails = set()
        for line in data:
            if line.startswith('ContactEmail:'):
                contact_emails = line.split('ContactEmail:')[1].strip()
                emails.update(contact_emails.split(','))
        return list(emails)

    def extract_test_program(self, data):
        for line in data:
            if line.startswith('TestProgram:'):
                return line.split('TestProgram:')[1].strip()
        return None

    def format_display_name(self, value):
        value = value.replace('[', ' [').replace(']', '] ')
        if len(value) > 6:
            value = value[:-6] + ' ' + value[-6:]
        if len(value) > 2:
            value = value[:-2] + ' ' + value[-2:]
        return value

    def create_table(self, data, header, table_name):
        if all(line.startswith('TestProgram:') or line.startswith('ContactEmail:') for line in data):
            return f"<p>There are no {table_name} submissions.</p>"

        table = "<table border='1' cellspacing='0' style='border-collapse: collapse;font-family: Calibri;'>\n"
        table += "  <tr>\n"
        for column in header:
            table += "      <th style='padding-right: 15px; text-align: left; background-color: #000000; color: #FFFFFF;'>{0}</th>\n".format(column.strip())
        table += "  </tr>\n"

        for row in data:
            if not row.startswith('TestProgram:') and not row.startswith('ContactEmail:'):
                columns = row.split('--')
                table += "  <tr style='margin: 0px; background-color: #DDEBF7;'>\n" if data.index(row) % 2 == 0 else "  <tr style='margin: 0px; background-color: #FFFFFF;'>\n"
                for column in columns:
                    parts = column.split(':')
                    if len(parts) == 2:
                        key, value = parts
                        if 'DisplayName' in key:
                            value = self.format_display_name(value)
                        table += "      <td style='padding-right: 15px;'>{0}</td>\n".format(value.strip())
                    else:
                        logging.warning(f"Unexpected format in line: {column}")
                table += "  </tr>\n"

        table += "</table>\n"
        return table

    def draft_email(self):
        mv_data = self.read_file(self.mv_table_file)
        correlation_data = self.read_file(self.correlation_table_file)

        if not mv_data and not correlation_data:
            logging.error("No data found in files. Email will not be drafted.")
            return

        # Use MV data for email addresses and test program (since both files have the same headers)
        email_addresses = self.extract_emails(mv_data if mv_data else correlation_data)
        test_program = self.extract_test_program(mv_data if mv_data else correlation_data)

        # Create tables with Ref_VPO positioned next to New_VPO
        correlation_header = ["DisplayName", "New_VPO", "Ref_VPO", "Lot", "Qty", "Operation", "EngID", "Tags"]
        mv_header = ["DisplayName", "New_VPO", "Ref_VPO", "Lot", "Qty", "Operation", "EngID", "Tags"]

        correlation_table = self.create_table(correlation_data, correlation_header, "Correlation") if correlation_data else "<p>There are no Correlation submissions.</p>"
        mv_table = self.create_table(mv_data, mv_header, "MV") if mv_data else "<p>There are no MV submissions.</p>"

        try:
            outlook = win32com.client.Dispatch('Outlook.Application')
            mail = outlook.CreateItem(0)

            mail.To = ','.join(email_addresses)
            mail.CC = "sam.ren.jie.yap@intel.com"
            mail.Subject = "VPO Submission Report"
            mail.HTMLBody = f"""
            <html style="font-family: Calibri; font-size: 0.9em;">
            <body>
            <p>Hello Team,</p>
            <p>Here is the VPO submission report. Please review the details below:</p>
            <p><strong>Correlation Submission</strong></p>
            <p><strong>Test Program:</strong> {test_program}</p>
            <p><strong>Email Recipients:</strong> {', '.join(email_addresses)}</p>
            {correlation_table}
            <p><strong>MV Submission</strong></p>
            <p><strong>Test Program:</strong> {test_program}</p>
            <p><strong>Email Recipients:</strong> {', '.join(email_addresses)}</p>
            {mv_table}
            <p>Thank you</p>
            </body></html>"""

            mail.Send()
            logging.info("Email drafted & sent successfully.")
        except Exception as e:
            logging.error(f"Error drafting email: {e}")


class SparkAutoSubmission:
    def __init__(self, excel_file_path, current_path, tp_site, common_path):
        self.excel_file_path = excel_file_path
        self.current_path = current_path
        self.tp_site = tp_site
        self.common_path = common_path

    def process(self):
        # Initialize SubmissionRead class and process the Excel file
        submission_read = SubmissionRead(self.excel_file_path, self.current_path)
        temp_file_path, vpo_run_directory, ssid = submission_read.process_excel()

        # Initialize SparkGenApi class and process submissions
        spark_gen_api = SparkGenApi(temp_file_path, vpo_run_directory, self.tp_site, self.common_path)
        spark_gen_api.process_submissions()

        # Initialize GenerateVpoJobsReport class and generate the report
        report_generator = GenerateVpoJobsReport(os.path.join(vpo_run_directory, 'submission_log.txt'), vpo_run_directory)
        report_generator.generate_report_info()

        # Wait for 12 minutes before starting GetVpoNumFromSpark
        wait_time = 720  # 12 minutes in seconds
        for remaining in range(wait_time, 0, -1):
            sys.stdout.write(f"\rWaiting for {remaining} seconds...")
            sys.stdout.flush()
            time.sleep(1)
        logging.info("\nWait complete. Proceeding with GetVpoNumFromSpark...")

        # Initialize GetVpoNumFromSpark class and run the GetData script
        report_file_path = os.path.join(vpo_run_directory, 'report_info.txt')
        submission_file_path = os.path.join(vpo_run_directory, 'submission_id.txt')  # Ensure this file is generated
        get_vpo_num_from_spark = GetVpoNumFromSpark(report_file_path, submission_file_path, vpo_run_directory, self.tp_site, self.common_path)
        get_vpo_num_from_spark.run_get_data_script()

        # Process output files (simplified)
        process_output = ProcessOutput(vpo_run_directory)
        process_output.process_output()

        # Draft and send email
        email_vpo_number = EmailVpoNumber(vpo_run_directory)
        email_vpo_number.draft_email()

        # Create timestamped ZIP archive
        self.create_archive(vpo_run_directory, ssid)

    def create_archive(self, vpo_run_directory, ssid):
        """Create timestamped ZIP archive and move all files"""
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        zip_directory = os.path.join(self.current_path, 'VPO_Description', 'VPO_Jobs', ssid, timestamp)
        os.makedirs(zip_directory, exist_ok=True)

        # Copy all files from vpo_run_directory to timestamped directory
        for file_name in os.listdir(vpo_run_directory):
            src_file = os.path.join(vpo_run_directory, file_name)
            if os.path.isfile(src_file):
                shutil.copy(src_file, zip_directory)
                logging.info(f"Copied {src_file} to {zip_directory}")

        # Create ZIP archive
        archive_path = shutil.make_archive(zip_directory, 'zip', zip_directory)
        logging.info(f"Created archive: {archive_path}")

        # Clean up the temporary directory after archiving
        shutil.rmtree(zip_directory)
        logging.info(f"Cleaned up temporary directory: {zip_directory}")


class SparkAutomationSelector:
    def __init__(self, excel_file_path, current_path, directory_path=None, tp_site=None, script_path=None):
        self.excel_file_path = excel_file_path
        self.current_path = current_path
        self.directory_path = directory_path
        self.tp_site = tp_site
        self.script_path = script_path
        self.common_path = script_path.replace('I:/', 'intel/') if script_path else None

    def run(self):
        # Use the processed common_path
        if self.directory_path:
            self.run_single_tp(self.common_path)
        else:
            # Check the Excel file for non-blank directory paths
            df_jobs = pd.read_excel(self.excel_file_path, sheet_name='VPO_Jobs')
            if df_jobs['DirectoryPath'].notna().any():
                self.run_multi_tp(self.common_path)
            else:
                raise ValueError("Error: No UserInput TP Path in both YAML file and Excel file. Please add to either one to proceed.")

    def run_single_tp(self, common_path):
        # Check for the required number of arguments
        if not self.directory_path or not self.excel_file_path or not self.current_path or not self.tp_site or not self.script_path:
            raise ValueError("Five arguments are required for SingleTP processing: tp_site, ExcelFilePath, current_path, script_path, and DirectoryPath.")

        # Process user input to determine TP_REV and update the Excel file
        user_input_processor = ProcessUserInput(self.directory_path)
        tp_rev = user_input_processor.determine_tp_rev()
        user_input_processor.update_excel(self.excel_file_path, self.directory_path, tp_rev)

        # Process the submission
        spark_auto_submission = SparkAutoSubmission(self.excel_file_path, self.current_path, self.tp_site, common_path)
        spark_auto_submission.process()

    def run_multi_tp(self, common_path):
        # Check for the required number of arguments
        if not self.excel_file_path or not self.current_path or not self.tp_site or not self.script_path:
            raise ValueError("Four arguments are required for MultiTP processing: tp_site, ExcelFilePath, current_path, and script_path.")

        # Process the Excel file to update TPRevision
        process_multi_tp_excel = ProcessMultiTPExcel(self.excel_file_path)
        process_multi_tp_excel.process_excel_multi()

        # Process the submission
        spark_auto_submission = SparkAutoSubmission(self.excel_file_path, self.current_path, self.tp_site, common_path)
        spark_auto_submission.process()


def main():
    # Get the tp_site, Excel file path, current path, script_path, and directory path from command-line arguments
    tp_site = sys.argv[1]
    excel_file_path = sys.argv[2]
    current_path = sys.argv[3]
    script_path = sys.argv[4]
    directory_path = sys.argv[5] if len(sys.argv) > 5 else None

    spark_automation_selector = SparkAutomationSelector(excel_file_path, current_path, directory_path, tp_site, script_path)
    spark_automation_selector.run()

    # Shut down logging to release file handles
    logging.shutdown()

    # Read SSID from Excel for cleanup
    df_static = pd.read_excel(excel_file_path, sheet_name='StaticInputs')
    ssid = df_static.at[0, 'SSID']
    
    # Clean up the vpo_run directory after archiving
    vpo_run_directory = os.path.join(current_path, 'Submission_History', ssid, 'vpo_run')
    if os.path.exists(vpo_run_directory):
        shutil.rmtree(vpo_run_directory)
        print(f"Cleaned up the vpo_run_directory: {vpo_run_directory}")


if __name__ == "__main__":
    main()
