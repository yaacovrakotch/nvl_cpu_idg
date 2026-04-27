import os
import sys
import subprocess
import json
import shutil
import re
import importlib
from datetime import datetime, timedelta
import logging
import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart


pd = None
load_workbook = None


def configure_proxy():
    proxy_url = 'http://proxy-dmz.intel.com:912'
    proxy_keys = ('http_proxy', 'https_proxy', 'HTTP_PROXY', 'HTTPS_PROXY')

    for key in proxy_keys:
        os.environ[key] = proxy_url

    for key in ('http_proxy', 'https_proxy'):
        subprocess.run(['setx', key, proxy_url], check=False, capture_output=True, text=True)


def ensure_runtime_dependencies():
    global pd, load_workbook

    required_packages = {
        'pandas': 'pandas',
        'openpyxl': 'openpyxl',
    }

    for module_name, package_name in required_packages.items():
        try:
            importlib.import_module(module_name)
        except ModuleNotFoundError:
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', package_name])

    import pandas as pandas_module
    from openpyxl import load_workbook as openpyxl_load_workbook

    pd = pandas_module
    load_workbook = openpyxl_load_workbook


def setup_logging(base_directory, debug_mode=False):
    # Configure logging to use the base directory for the log file
    log_file_path = os.path.join(base_directory, 'submission_log.txt')

    # Clear any existing handlers to avoid conflicts
    for handler in logging.root.handlers[:]:
        logging.root.removeHandler(handler)

    # Set log level based on debug mode
    log_level = logging.DEBUG if debug_mode else logging.INFO

    # Configure logging
    logging.basicConfig(
        filename=log_file_path,
        level=log_level,
        format='%(asctime)s - %(levelname)s - %(message)s',
        filemode='w'  # Overwrite the file for each site
    )

    # Only add console handler in debug mode
    if debug_mode:
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        console_handler.setFormatter(formatter)
        logging.getLogger().addHandler(console_handler)

    logging.info(f"Logging initialized. Log file path: {log_file_path}")


class EmailValidator:
    def __init__(self):
        self.admin_emails = {'sam.ren.jie.yap@intel.com', 'tai.pham@intel.com'}

    def validate_contact_emails_from_logs(self, all_site_data):
        """
        Validate that contact emails from log files are not admin emails only
        """
        validation_errors = []

        for site_data in all_site_data:
            tp_site = site_data['tp_site']
            log_file_path = os.path.join(site_data['vpo_run_directory'], 'submission_log.txt')

            try:
                # Read the log file and extract contact emails
                contact_emails_set = set()

                with open(log_file_path, 'r', encoding='utf-8') as file:
                    for line in file:
                        line = line.strip()
                        # Look for JSON data in log entries
                        if "INFO - {" in line:
                            # Extract the JSON string from the log entry
                            json_start = line.find("INFO - {")
                            if json_start != -1:
                                json_str = line[json_start + 7:]  # Skip "INFO - "
                                try:
                                    # Parse JSON safely
                                    entry = json.loads(json_str)
                                    # Extract emails from the Contact field
                                    contact_emails = entry.get('Contact', '')
                                    if contact_emails:
                                        contact_emails_set.update(email.strip().lower() for email in str(contact_emails).split(','))
                                except json.JSONDecodeError:
                                    continue

                # Validate the extracted emails
                if not contact_emails_set:
                    validation_errors.append(f"TPSite '{tp_site}': No contact emails found in log data")
                    continue

                # Check if all emails are admin emails
                non_admin_emails = [email for email in contact_emails_set if email not in self.admin_emails]

                if not non_admin_emails:
                    validation_errors.append(f"TPSite '{tp_site}': Only admin emails found ({', '.join(contact_emails_set)})")

            except Exception as e:
                validation_errors.append(f"TPSite '{tp_site}': Error reading log file - {str(e)}")

        return validation_errors

    def prompt_for_email_correction(self, validation_errors):
        """
        Display validation errors and prompt user to fix them
        """
        print("\n" + "="*80)
        print("EMAIL VALIDATION FAILED")
        print("="*80)
        print("The following issues were found with contact emails:")
        print()

        for i, error in enumerate(validation_errors, 1):
            print(f"{i}. {error}")

        print()
        print("REQUIRED ACTION:")
        print("Please input the owner of the VPO email or emails by comma separated")
        print()
        print("Examples of valid email formats:")
        print("  - sam.ren.jie.yap@intel.com")
        print("  - sam.ren.jie.yap@intel.com, tai.pham@intel.com")
        print()
        print("Note: Admin emails (sam.ren.jie.yap@intel.com, tai.pham@intel.com) can be")
        print("      included but should not be the only contacts.")
        print()
        print("Please update the 'Contact' column in the 'StaticInputs' sheet of your Excel file")
        print("and run the script again.")
        print("="*80)

        # Exit the script
        sys.exit(1)


class SiteProcessingStatus:
    def __init__(self):
        self.successful_sites = []
        self.failed_sites = []

    def add_success(self, site, directory):
        self.successful_sites.append({'site': site, 'directory': directory})

    def add_failure(self, site, error):
        self.failed_sites.append({'site': site, 'error': str(error)})

    def get_summary(self):
        return {
            'successful_count': len(self.successful_sites),
            'failed_count': len(self.failed_sites),
            'successful_sites': [s['site'] for s in self.successful_sites],
            'failed_sites': [f['site'] for f in self.failed_sites]
        }


class ProcessMultiTPExcel:
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path

    def update_excel_multi(self, directory_path, tp_rev):
        # Print the user input path
        print(f"Processing Directory_Path: {directory_path}")

        # Load the workbook and select the sheet
        workbook = load_workbook(self.excel_file_path)
        sheet = workbook['VPO_Jobs']

        # Iterate over the rows and update the TPRevision column
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            if row[1].value == directory_path:  # Check if the DirectoryPath matches
                additional_tags = row[5].value  # AdditionalTags is now the sixth column (index 5) after adding TPSite
                if additional_tags and additional_tags.strip():  # Check if AdditionalTags is not blank
                    row[2].value = f"{tp_rev}_{additional_tags.strip()}"  # Update TPRevision
                else:
                    row[2].value = tp_rev  # Update TPRevision without AdditionalTags

        # Save the workbook
        workbook.save(self.excel_file_path)

    def process_excel_multi(self):
        # Read the Excel file to get the 'VPO_Jobs' sheet
        df_jobs = pd.read_excel(self.excel_file_path, sheet_name='VPO_Jobs')

        # Iterate over each row to process the directory path and determine TP_REV
        for index, row in df_jobs.iterrows():
            directory_path = row['DirectoryPath']

            # Process user input to determine TP_REV
            user_input_processor = ProcessUserInput(directory_path)
            try:
                tp_rev = user_input_processor.determine_tp_rev()
                # Update the Excel file with the determined TP_REV
                self.update_excel_multi(directory_path, tp_rev)
                print(f"Updated TP_REV for directory {directory_path}: {tp_rev}")
            except ValueError as e:
                print(f"Error determining TP_REV for directory {directory_path}: {e}")
                continue


class ProcessUserInput:
    def __init__(self, directory_path):
        self.directory_path = directory_path

    def determine_tp_rev(self):
        # Check if the directory exists
        if os.path.isdir(self.directory_path):
            print(f"Directory exists: {self.directory_path}")
        else:
            raise ValueError(f"Directory does not exist: {self.directory_path}")

        # Define possible paths to the usrv files
        paths = [
            os.path.join(self.directory_path, 'Shared', 'BaseInputs', 'UservarDefinitions.usrv'),
            os.path.join(self.directory_path, 'Shared', 'BaseInputs', 'Common', 'Common_Files', 'common.usrv')
        ]

        # Search for TP_REV in the files
        for path in paths:
            if os.path.exists(path):
                print(f"File exists: {path}")  # Print the file path if it exists
                with open(path, 'r') as file:
                    for line in file:
                        if 'Const String TPName' in line:
                            # Extract TP_REV from the line
                            tp_rev = line.split('"')[1]
                            print(f"Full TP_REV before trim: {tp_rev}")  # Print the full TP_REV before trimming
                            # Extract the desired substring from TP_REV
                            trimmed_tp_rev = tp_rev[10:15]  # Adjust indices based on your requirement
                            print(f"Trimmed TP_REV: {trimmed_tp_rev}")  # Print the trimmed TP_REV
                            return trimmed_tp_rev
            else:
                print(f"File does not exist: {path}")  # Print a warning if the file does not exist

        raise ValueError("TP_REV not found in the specified files.")

    def update_excel(self, excel_file_path, directory_path, tp_rev):
        # Print the user input path
        print(f"User input Directory_Path: {directory_path}")

        # Load the workbook and select the sheet
        workbook = load_workbook(excel_file_path)
        sheet = workbook['VPO_Jobs']

        # Iterate over the rows and update the TPRevision column
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            additional_tags = row[5].value  # AdditionalTags is now the sixth column (index 5) after adding TPSite
            if additional_tags and additional_tags.strip():  # Check if AdditionalTags is not blank
                row[2].value = f"{tp_rev}_{additional_tags.strip()}"  # Update TPRevision
            else:
                row[2].value = tp_rev  # Update TPRevision without AdditionalTags

        # Update the DirectoryPath column
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            row[1].value = directory_path  # Assuming 'DirectoryPath' is the second column

        # Save the workbook
        workbook.save(excel_file_path)


class SubmissionRead:
    def __init__(self, excel_file_path, current_path):
        self.excel_file_path = self.convert_path(excel_file_path)
        self.current_path = current_path

    def convert_path(self, path):
        # Convert backslashes to forward slashes for cross-platform compatibility
        return path.replace('\\', '/')

    def normalize_path(self, path):
        """Normalize path to use OS-appropriate separators"""
        return os.path.normpath(path.replace('\\', os.sep).replace('/', os.sep))

    def read_excel_file(self, file_path, sheet_name):
        # Read the specified sheet from the Excel file
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        print(df.to_string())
        return df

    def find_files(self, base_path, tp_type):
        # Generic regular expressions to match the desired file patterns
        stpl_pattern = re.compile(r'.*SubTestPlan.*\.stpl$')
        plist_pattern = re.compile(r'.*PLIST_ALL_CLASS_.*\.plist\.xml$')

        stpl_path = None
        plist_path = None

        search_path = os.path.join(base_path, tp_type)

        for root, dirs, files in os.walk(search_path):
            for file in files:
                full_path = os.path.join(root, file)
                full_path = self.convert_path(full_path)  # Ensure path uses forward slashes
                if stpl_pattern.match(full_path):
                    stpl_path = full_path.replace(base_path + '/', '')
                elif plist_pattern.match(full_path):
                    plist_path = full_path.replace(base_path + '/', '')

        if not stpl_path or not plist_path:
            raise ValueError("Could not find STPL or PLIST files with the specified patterns.")

        return stpl_path, plist_path

    def determine_tp_type(self, directory_path):
        # Automatically determine TP_TYPE by checking the directory structure
        for root, dirs, files in os.walk(directory_path):
            if 'POR_TP' in dirs:
                return 'POR_TP'
            elif 'ENG_TP' in dirs:
                return 'ENG_TP'
        raise ValueError("Could not determine TP_TYPE. Please ensure the directory contains either 'POR_TP' or 'ENG_TP'.")

    def extract_excel_inputs(self, df_jobs, df_static, df_submission_option, df_site_labs, index):
        # Extract values from the VPO_Jobs sheet
        Directory_Path = df_jobs.at[index, 'DirectoryPath']
        TP_REV = df_jobs.at[index, 'TPRevision']
        TPSite = df_jobs.at[index, 'TPSite']  # Get TPSite from VPO_Jobs row

        # Extract Ref_VPO and handle NaN values properly
        Ref_VPO = df_jobs.at[index, 'Ref_VPO']
        if pd.isna(Ref_VPO) or Ref_VPO == '' or str(Ref_VPO).strip() == '' or str(Ref_VPO).lower() == 'nan':
            Ref_VPO = "N/A"  # Set default value instead of nan
        else:
            Ref_VPO = str(Ref_VPO).strip()  # Convert to string and strip whitespace

        AdditionalTags = df_jobs.at[index, 'AddtionalTags']  # Note: keeping the typo as it appears in your column header
        OverwritingPartType = df_jobs.at[index, 'OverwritingPartType']
        FusedPart = df_jobs.at[index, 'FusedPart']
        QDFS = df_jobs.at[index, 'QDFS']
        hri = df_jobs.at[index, 'hri']
        mrv = df_jobs.at[index, 'mrv']
        recipe_config = df_jobs.at[index, 'recipe_config']
        print(f'Recipe config: {recipe_config}')
        SourceLot = df_jobs.at[index, 'SourceLot']
        VisualIDPath = df_jobs.at[index, 'VisualIDPath']
        NumberofUnitsToRun = df_jobs.at[index, 'NumberofUnitsToRun']
        Stepping = df_jobs.at[index, 'Stepping']
        PartType = df_jobs.at[index, 'PartType']
        Location = df_jobs.at[index, 'Location']
        EngID = df_jobs.at[index, 'EngID']
        AdditionalComments = df_jobs.at[index, 'AdditionalComments']  # Extract AdditionalComments
        SubmissionOption = df_jobs.at[index, 'SubmissionOption']

        # Extract values from the StaticInputs sheet based on TPSite
        static_row = df_static.loc[df_static['TPSite'] == TPSite]
        if static_row.empty:
            raise ValueError(f"No StaticInputs found for TPSite: {TPSite}")

        SSID = static_row['SSID'].iloc[0]
        Contact = static_row['Contact'].iloc[0]

        # Get SiteChosen from StaticInputs sheet
        if 'SiteChosen' in static_row.columns:
            SiteChosen = static_row['SiteChosen'].iloc[0]
            print(f"Using SiteChosen '{SiteChosen}' from StaticInputs for TPSite '{TPSite}'")
        else:
            # Fallback: use TPSite as SiteChosen
            SiteChosen = TPSite
            print(f"Warning: SiteChosen column not found in StaticInputs, using TPSite '{TPSite}' as SiteChosen")

        # Extract TAG and SpecialInstruction from the SubmissionOption sheet
        submission_info = df_submission_option.loc[df_submission_option['SubmissionOption'] == SubmissionOption]
        if submission_info.empty:
            raise ValueError(f"No SubmissionOption found for: {SubmissionOption}")
        TAG = submission_info['TAG'].values[0]
        SpecialInstruction = submission_info['SpecialInstruction'].values[0]

        # Extract segment, team, and lab information from the Site-Labs sheet using SiteChosen
        site_info = df_site_labs.loc[df_site_labs['Selector'] == SiteChosen]
        if site_info.empty:
            # Print available selectors for debugging
            available_selectors = df_site_labs['Selector'].unique().tolist()
            print(f"Available Selectors in Site-Labs: {available_selectors}")
            raise ValueError(f"No Site-Labs info found for SiteChosen: {SiteChosen}. Available selectors: {available_selectors}")

        Segment = site_info['Segment'].iloc[0]
        Team = site_info['Team'].iloc[0]
        Labs = site_info['Labs'].iloc[0]

        # Strip whitespace from recipe_config
        recipe_config = str(recipe_config).strip()

        # Validate and process inputs
        if not os.path.isdir(Directory_Path):
            raise ValueError("Path is Invalid!")
        BaseTP = self.convert_path(Directory_Path)

        # Automatically determine TP_TYPE
        TP_TYPE = self.determine_tp_type(Directory_Path)

        if OverwritingPartType == "Y":
            OveridePartType = "True"
        elif OverwritingPartType == "N":
            OveridePartType = "False"
        else:
            raise ValueError("Invalid OverwritingPartType value. Please use 'Y' or 'N'.")

        if FusedPart == "N":
            QDFS = "XXXX"

        if recipe_config == "1":
            recipe = "1"
        elif recipe_config == "2":
            recipe = "2"
        else:
            raise ValueError("Invalid recipe_config value. Please use '1' or '2'.")

        # Set default values for hri and mrv if specified
        if hri == "Default":
            hri = "DEFAULT_HRI"
        if mrv == "Default":
            mrv = "000000000000"

        return (BaseTP, Segment, Team, TP_TYPE, TP_REV, OveridePartType, QDFS, hri, mrv, recipe,
                SourceLot, VisualIDPath, NumberofUnitsToRun, Stepping, PartType, Location, EngID, TAG, SpecialInstruction, Contact, SSID, SubmissionOption, Labs, AdditionalComments, Ref_VPO, TPSite, AdditionalTags)

    def process_excel(self):
        # Process the Excel file and create a temporary file with processed data
        if not os.path.isfile(self.excel_file_path):
            print("Invalid file path. Please ensure the file exists.")
            return

        df_jobs = self.read_excel_file(self.excel_file_path, sheet_name='VPO_Jobs')
        df_static = self.read_excel_file(self.excel_file_path, sheet_name='StaticInputs')
        df_site_labs = self.read_excel_file(self.excel_file_path, sheet_name='Site-Labs')
        df_submission_option = self.read_excel_file(self.excel_file_path, sheet_name='SubmissionOption')

        # Group by TPSite to create separate directories
        site_groups = df_jobs.groupby('TPSite')

        all_data_entries = []

        for tp_site, site_df in site_groups:
            print(f"\n=== Processing TPSite: {tp_site} ===")

            # Get SSID from StaticInputs for this TPSite
            static_row = df_static.loc[df_static['TPSite'] == tp_site]
            if static_row.empty:
                print(f"Warning: No StaticInputs found for TPSite: {tp_site}")
                continue

            ssid = static_row['SSID'].iloc[0]

            # Create site-specific directory
            sub_dielet = os.environ.get('DIELET', '')
            if sub_dielet:
                vpo_run_directory = os.path.join(self.current_path, 'Submission_History', ssid, tp_site, sub_dielet, 'vpo_run')
            else:
                vpo_run_directory = os.path.join(self.current_path, 'Submission_History', ssid, tp_site, 'vpo_run')

            # Normalize the path to use the OS-appropriate separators
            vpo_run_directory = self.normalize_path(vpo_run_directory)

            if os.path.exists(vpo_run_directory):
                shutil.rmtree(vpo_run_directory)

            os.makedirs(vpo_run_directory, exist_ok=True)
            print(f"Created directory: {vpo_run_directory}")

            # Set up logging for this site
            setup_logging(vpo_run_directory)

            # Open the temporary file for writing
            temp_file_path = os.path.join(vpo_run_directory, 'vpo_jobs.temp')
            logging.info(f"Creating temporary file at: {temp_file_path}")

            with open(temp_file_path, 'w') as temp_file:
                # Iterate over each row in this site's DataFrame
                for index in site_df.index:
                    try:
                        inputs = self.extract_excel_inputs(df_jobs, df_static, df_submission_option, df_site_labs, index)

                        # Find STPL and PLIST paths using regular expressions
                        STPL, PLIST = self.find_files(inputs[0], inputs[3])  # BaseTP and TP_TYPE

                        # Extract grp_displayname_tprev from TP_REV
                        grp_displayname_tprev = inputs[4][:5]  # TP_REV is the fifth element in inputs

                        # Log extracted values (simplified format for parsing)
                        log_data = {
                            "BaseTP": inputs[0],
                            "Segment": inputs[1],
                            "Team": inputs[2],
                            "TP_TYPE": inputs[3],
                            "STPL": STPL,
                            "PLIST": PLIST,
                            "TP_REV": inputs[4],
                            "OveridePartType": inputs[5],
                            "QDFS": inputs[6],
                            "hri": inputs[7],
                            "mrv": inputs[8],
                            "recipe": inputs[9],
                            "SourceLot": inputs[10],
                            "VisualIDPath": inputs[11],
                            "NumberofUnitsToRun": inputs[12],
                            "Stepping": inputs[13],
                            "PartType": inputs[14],
                            "Location": inputs[15],
                            "EngID": inputs[16],
                            "TAG": inputs[17],
                            "SpecialInstruction": inputs[18],
                            "Contact": inputs[19],
                            "SSID": inputs[20],
                            "SubmissionOption": inputs[21],
                            "Labs": inputs[22],
                            "AdditionalComments": inputs[23],
                            "Ref_VPO": inputs[24],
                            "TPSite": inputs[25],
                            "AdditionalTags": inputs[26],
                            "grp_displayname_tprev": grp_displayname_tprev
                        }

                        # Convert to JSON string for safe logging (avoids nan issues)
                        log_json = json.dumps(log_data, default=str)
                        logging.info(log_json)
                        print(log_data)
                        # Write to temp file
                        temp_file.write(f"Processed values for row {index}:\n")
                        for input_name, input_value in zip(["BaseTP", "Segment", "Team", "TP_TYPE", "STPL", "PLIST", "TP_REV", "OveridePartType", "QDFS", "hri", "mrv", "recipe",
                                                            "SourceLot", "VisualIDPath", "NumberofUnitsToRun", "Stepping", "PartType", "Location", "EngID", "TAG", "SpecialInstruction", "Contact", "SSID", "SubmissionOption", "Labs", "AdditionalComments", "Ref_VPO", "TPSite", "AdditionalTags", "grp_displayname_tprev"], 
                                                           [*inputs[:4], STPL, PLIST, *inputs[4:], grp_displayname_tprev]):
                            temp_file.write(f"{input_name}: {input_value}\n")
                        temp_file.write("\n")  # Add a newline between entries

                    except ValueError as e:
                        logging.error(f"Error processing row {index}: {e}")
                        print(f"Error processing row {index} for site {tp_site}: {e}")

            # Copy the Excel file to the vpo_run directory
            shutil.copy(self.excel_file_path, vpo_run_directory)

            # Store data for this site including expected count
            all_data_entries.append({
                'temp_file_path': temp_file_path,
                'vpo_run_directory': vpo_run_directory,
                'ssid': ssid,
                'tp_site': tp_site,
                'expected_vpo_count': len(site_df)  # Track expected number of VPOs for this site
            })

        print(f"\n=== Processed {len(all_data_entries)} sites total ===")
        # Return all site data
        return all_data_entries


class GenerateVpoJobsReport:
    def __init__(self, log_file_path, vpo_run_directory):
        self.log_file_path = log_file_path
        self.vpo_run_directory = vpo_run_directory

    def read_log_file(self):
        # Read the log file and parse the data into a list of dictionaries
        data_entries = []
        contact_emails_set = set()

        try:
            with open(self.log_file_path, 'r', encoding='utf-8') as file:
                for line in file:
                    line = line.strip()
                    # Look for JSON data in log entries
                    if "INFO - {" in line:
                        # Extract the JSON string from the log entry
                        json_start = line.find("INFO - {")
                        if json_start != -1:
                            json_str = line[json_start + 7:]  # Skip "INFO - "
                            try:
                                # Parse JSON safely
                                entry = json.loads(json_str)
                                data_entries.append(entry)
                                # Extract emails from the Contact field
                                contact_emails = entry.get('Contact', '')
                                if contact_emails:
                                    contact_emails_set.update(email.strip() for email in str(contact_emails).split(','))
                            except json.JSONDecodeError as e:
                                logging.debug(f"JSON decode error: {e}")
                                continue
        except FileNotFoundError:
            logging.error(f"Log file not found: {self.log_file_path}")
            raise
        except Exception as e:
            logging.error(f"Error reading log file: {e}")
            raise

        logging.info(f"Extracted {len(data_entries)} data entries")
        return data_entries, contact_emails_set

    def generate_report_info(self):
        # Read data from the log file
        data_entries, contact_emails_set = self.read_log_file()

        # Check if data_entries is empty or missing expected keys
        if not data_entries or 'BaseTP' not in data_entries[0]:
            logging.error("Missing 'BaseTP' in data entries. Please check the log file format.")
            raise KeyError("Missing 'BaseTP' in data entries. Please check the log file format.")

        # Extract TP name from BaseTP
        base_tp = data_entries[0]['BaseTP']
        logging.info(f"BaseTP value is {base_tp}")

        # Extract TP name using string manipulation
        tp_name = base_tp.split('/')[-1]  # Get the last segment after the last '/'

        # Generate report
        report_path = os.path.join(self.vpo_run_directory, 'report_info.txt')

        with open(report_path, 'w') as report_file:
            # Write TestProgram and ContactEmail to both files
            report_file.write(f'TestProgram: {tp_name}\n')
            report_file.write(f'ContactEmail: {",".join(contact_emails_set)}\n')

            # Initialize a counter dictionary for each SubmissionOption
            counters = {}

            # Sort entries by SubmissionOption to ensure correct ordering
            sorted_entries = sorted(data_entries, key=lambda x: x['SubmissionOption'])

            for entry in sorted_entries:
                submission_option = entry['SubmissionOption']
                if submission_option not in counters:
                    counters[submission_option] = 0  # Initialize counter for each SubmissionOption

                counters[submission_option] += 1  # Increment counter for the current SubmissionOption

                # Extract tag_prefix from PLIST path
                plist_segments = entry['PLIST'].split('/')
                if len(plist_segments) > 2:
                    tag_prefix = plist_segments[2].split('PLIST_ALL_')[1].split('.')[0]
                else:
                    raise ValueError("PLIST path does not have the expected format.")

                # Get DIELET from environment variable
                dielet = os.environ.get('DIELET', '')

                display_name = f"{submission_option} [{counters[submission_option]:02}] {entry['TP_REV']} {entry['Location']} {entry['EngID']}"  # Construct DisplayName
                ref_vpo = entry['Ref_VPO']  # Get Ref_VPO from the entry
                source_lot = entry['SourceLot']
                num_units = entry['NumberofUnitsToRun']
                location = entry['Location']
                eng_id = entry['EngID']

                # Create tags for report (show all four tags separated by space)
                additional_tags = entry.get('AdditionalTags', '')
                if dielet:
                    if additional_tags and str(additional_tags).strip() and str(additional_tags).strip().lower() != 'nan':
                        tag = f"{tag_prefix} {dielet}_{entry['TP_REV']} {entry['TAG']} {str(additional_tags).strip()}"
                    else:
                        tag = f"{tag_prefix} {dielet}_{entry['TP_REV']} {entry['TAG']}"
                else:
                    # Fallback if DIELET is not set
                    if additional_tags and str(additional_tags).strip() and str(additional_tags).strip().lower() != 'nan':
                        tag = f"{tag_prefix} {entry['TP_REV']} {entry['TAG']} {str(additional_tags).strip()}"
                    else:
                        tag = f"{tag_prefix} {entry['TP_REV']} {entry['TAG']}"

                # Write in the format expected by GetData.ps1: DisplayName-Lot-Qty-Operation-EngID-Tags-Ref_VPO
                report_file.write(f'{display_name}-{source_lot}-{num_units}-{location}-{eng_id}-{tag}-{ref_vpo}\n')

        logging.info(f"Report generated at: {report_path}")


class SparkGenApi:
    def __init__(self, temp_file_path, vpo_run_directory, tp_site_excel, script_path):
        self.temp_file_path = temp_file_path
        self.vpo_run_directory = vpo_run_directory
        self.tp_site_excel = tp_site_excel  # TPSite from Excel (PG7, PG12, JF)
        self.script_path = script_path

        # Configurable fields (can be set via environment variables)
        self.expected_bin = os.environ.get('EXPECTED_BIN', 'Bin01')
        self.proceed_with_remaining = os.environ.get('PROCEED_WITH_REMAINING', 'Yes')
        self.eta_hours_offset = int(os.environ.get('ETA_HOURS_OFFSET', '8'))

    def read_temp_file(self):
        # Check if the temporary file exists before reading
        if not os.path.exists(self.temp_file_path):
            logging.error(f"Temporary file not found: {self.temp_file_path}")
            raise FileNotFoundError(f"Temporary file not found: {self.temp_file_path}")

        # Read the temporary file and parse the data into a list of dictionaries
        data = []
        with open(self.temp_file_path, 'r') as file:
            entry = {}
            for line in file:
                line = line.strip()
                if line.startswith("Processed values for row"):
                    if entry:
                        data.append(entry)
                    entry = {}
                elif line:
                    if ": " in line:
                        key, value = line.split(": ", 1)
                        entry[key] = value

            if entry:
                data.append(entry)

        logging.debug("Data entries read from temp file:")
        for entry in data:
            logging.debug(entry)

        return data

    def convert_base_tp_to_network_path(self, base_tp):
        # Convert the base TP path to a network path based on the Excel TPSite
        # the function here help to use submission by the site input
        # TP need to resides trigger from either machine the script will check based on mapped site Idrive
        # comes to this function, it will help to convert based on the excel tp_site inputs
        if 'I:/' in base_tp:
            if self.tp_site_excel.upper() in ['PG7', 'PG12']:
                base_tp = base_tp.replace('I:/', '\\\\gar.corp.intel.com\\ec\\proj\\mdl\\pg\\intel\\')
            elif self.tp_site_excel.upper() == 'JF':
                base_tp = base_tp.replace('I:/', '\\\\amr.corp.intel.com\\ec\\proj\\mdl\\jf\\intel\\')
            elif self.tp_site_excel.upper() == 'FM':
                base_tp = base_tp.replace('I:/', '\\\\amr.corp.intel.com\\ec\\proj\\mdl\\fm\\intel\\')
            elif self.tp_site_excel.upper() == 'IDC':
                base_tp = base_tp.replace('I:/', '\\\\ger.corp.intel.com\\ec\\proj\\mdl\\ha\\intel\\')
            elif self.tp_site_excel.upper() == 'BA':
                base_tp = base_tp.replace('I:/', '\\\\gar.corp.intel.com\\ec\\proj\\mdl\\ba\\intel\\')
            else:
                raise ValueError(f"Invalid Excel TPSite value: {self.tp_site_excel}. Please use 'PG7', 'PG12', 'JF', 'FM', 'IDC', or 'BA'.")
            base_tp = base_tp.replace('/', '\\')
        return base_tp

    def read_visual_ids(self, visual_id_path):
        # Read the Visual IDs from the specified file path
        with open(visual_id_path, 'r') as file:
            data = [line.strip() for line in file.readlines()]
        return data

    def remove_none_values(self, d):
        """Recursively remove None values from dictionaries."""
        if isinstance(d, dict):
            return {k: self.remove_none_values(v) for k, v in d.items() if v is not None}
        elif isinstance(d, list):
            return [self.remove_none_values(v) for v in d if v is not None]
        else:
            return d

    def calculate_eta(self):
        """Calculate ETA (current time + offset hours) in YYWWDD format with Day/Night suffix"""
        eta_date = datetime.now() + timedelta(hours=self.eta_hours_offset)

        # Format as YYWWDD (YY = year, WW = week number, DD = day of week)
        year = eta_date.strftime('%y')
        week = eta_date.strftime('%U')  # Week number (00-53)

        # Day of week: Monday=01, Tuesday=02, ..., Saturday=06, Sunday=00
        day_of_week = eta_date.isoweekday() % 7  # This converts 7 (Sunday) to 0

        # Determine Day or Night based on hour (24-hour format)
        hour = eta_date.hour
        if 6 <= hour < 18:  # 6:00 AM to 5:59 PM
            shift = "Day"
        else:  # 6:00 PM to 5:59 AM
            shift = "Night"

        eta_formatted = f"{year}{week:0>2}{day_of_week:02d} {shift}"

        return eta_formatted

    def create_submission_json(self, data_entries, submission_type):
        # Create a JSON file for the submission based on the data entries and submission type
        submission_file = os.path.join(self.vpo_run_directory, f'submission_vpo_{self.tp_site_excel}_{submission_type}.json')

        # Remove old submission file if exists
        if os.path.exists(submission_file):
            os.remove(submission_file)

        experiments = []
        counter = 0  # Initialize counter for each submission type
        for data_entry in data_entries:
            counter += 1
            # Convert BaseTP to network path using Excel TPSite
            base_tp_network_path = self.convert_base_tp_to_network_path(data_entry['BaseTP'])

            # Prepare the stplDirectory
            path_stpl = data_entry['STPL'].split('/')[0] + '\\' + data_entry['STPL'].split('/')[1]
            stpl_directory = base_tp_network_path + '\\' + path_stpl

            # Check if VisualIDPath is valid
            use_vid = False
            visual_ids = []
            visual_id_path = data_entry.get('VisualIDPath', None)
            if visual_id_path and pd.notna(visual_id_path) and os.path.exists(visual_id_path):
                use_vid = True
                visual_ids = self.read_visual_ids(visual_id_path)

            # Determine recipe mode
            recipe_mode = data_entry['recipe']
            recipe_directory = os.path.join(base_tp_network_path, 'astra') if recipe_mode == "2" else None

            # Extract the desired portion from the PLIST path
            plist_segments = data_entry['PLIST'].split('/')
            if len(plist_segments) > 2:
                # Extract the part after 'PLIST_ALL_' in the third segment
                tag_prefix = plist_segments[2].split('PLIST_ALL_')[1].split('.')[0]
            else:
                raise ValueError("PLIST path does not have the expected format.")

            # Get DIELET from environment variable
            dielet = os.environ.get('DIELET', '')

            # Create four separate tags
            tags_list = [tag_prefix]  # Tag1

            # Tag2
            if dielet:
                tags_list.append(f"{dielet}_{data_entry['TP_REV']}")
            else:
                tags_list.append(data_entry['TP_REV'])  # Fallback if DIELET is not set

            # Tag3
            tags_list.append(data_entry['TAG'])

            # Tag4 - Add AdditionalTags if it exists and is not empty
            additional_tags = data_entry.get('AdditionalTags', '')
            if additional_tags and str(additional_tags).strip() and str(additional_tags).strip().lower() != 'nan':
                tags_list.append(str(additional_tags).strip())

            # Get ContactEmails
            contact_data = data_entry.get('Contact', '')
            contact_emails = []
            if contact_data and str(contact_data).strip() and str(contact_data).strip().lower() != 'nan':
                for email in str(contact_data).strip().split(','):
                    contact_emails.append(email)
                print(contact_emails)

            # Set experimentType and activityType based on submission type
            if submission_type in ["MV", "MVPFused"]:
                experiment_type = "Engineering"
                activity_type = "ModuleValidation"
            elif submission_type in ["Correlation", "PreFuse", "PostFuse"]:
                experiment_type = "Correlation"
                activity_type = "Correlation"
            else:
                # Default fallback for any other submission types
                experiment_type = "Correlation"
                activity_type = "Correlation"

            # Calculate ETA
            eta_formatted = self.calculate_eta()

            # Create the new formatted comment
            formatted_comment = f"""For any issue, please contact:
PDE name: {contact_data if contact_data and str(contact_data).strip() and str(contact_data).strip().lower() != 'nan' else 'N/A'}
ETA: {eta_formatted}
Expected Bin: {self.expected_bin}
Proceed with remaining units (Yes/No): {self.proceed_with_remaining}

{data_entry['SpecialInstruction']}
{data_entry.get('AdditionalComments', '')}"""

            # Prepare the experiment data
            experiment = {
                "displayName": f"{submission_type} [{counter:02}] {data_entry['TP_REV']} {data_entry['Location']} {data_entry['EngID']}",
                "experimentType": experiment_type,
                "activityType": activity_type,
                "bomGroupName": tag_prefix,
                "step": data_entry['Stepping'],
                "tplFileName": None if recipe_mode == "2" else "..\\..\\BaseTestPlan.tpl",
                "stplFileName": None if recipe_mode == "2" else data_entry['STPL'].split('/')[2],
                "environmentFileName": "EnvironmentFile.env" if submission_type == "PostFuse" or recipe_mode != "2" else None,
                "plistAllFileName": None if recipe_mode == "2" else data_entry['PLIST'].split('/')[2],
                "testTimeInSecPerUnit": 400,
                "retestRate": 50,
                "lab": data_entry['Labs'],
                "material": {
                    "materialIssue": {
                        "materialIssueIsRequired": False
                    },
                    "units": [{"visualId": vid} for vid in visual_ids] if use_vid else None,
                    "lots": [
                        {
                            "name": data_entry['SourceLot']
                        }
                        if int(data_entry['NumberofUnitsToRun']) == 0 else {
                            "name": data_entry['SourceLot'],
                            "numOfUnitsToRun": int(data_entry['NumberofUnitsToRun'])
                        }
                    ] if not use_vid else None,
                    "partTypeOverride": data_entry['PartType'] if data_entry['OveridePartType'] == "True" else None,
                    "sspec": data_entry['QDFS'] if submission_type == "PostFuse" or submission_type == "MVPFused" else None
                },
                "conditions": [
                    {
                        "operation": data_entry['Location'],
                        "engineeringId": data_entry['EngID'],
                        "dieSelection": "NA",
                        "comment": formatted_comment,  # Use the new formatted comment
                        "moveUnits": "Good"
                    }
                ],
                "flexbom": {
                    "hri": data_entry['hri'],
                    "mrv": data_entry['mrv']
                },
                "experimentState": "Ready",
                "tags": tags_list,
                "ContactEmails": contact_emails,
                "recipe": {
                    "recipeSource": "ByoUseAsIs" if recipe_mode == "2" else "TpGenerated",
                    "recipePath": recipe_directory if recipe_mode == "2" else None
                }
            }

            experiments.append(experiment)

        grp_displayname_tprev = data_entries[0]['grp_displayname_tprev']

        # Prepare the JSON data
        json_data = {
            "segment": data_entries[0]['Segment'],
            "team": data_entries[0]['Team'],
            "displayName": f"{self.tp_site_excel} {submission_type} {grp_displayname_tprev}",
            "stplDirectory": stpl_directory,
            "experiments": experiments
        }

        # Remove None values from the JSON data
        json_data_cleaned = self.remove_none_values(json_data)

        # Write the cleaned JSON data to the file
        with open(submission_file, 'w') as f:
            json.dump(json_data_cleaned, f, indent=4)

        logging.info(f"{self.tp_site_excel}_{submission_type} submission JSON is successfully created.")

    # TODO: update the ps file to have multiple account selection based on site
    def submit_to_spark(self, submission_file):

        # Call the PowerShell script to submit the JSON file to SPARK
        logging.info(f"Calling the Spark submission file for Excel TPSite {self.tp_site_excel}...")
        output_directory = self.vpo_run_directory

        # Use site prefix based on Excel TPSite
        script_file = os.path.normpath(os.path.join(self.script_path, "submission_MV.ps1"))

        try:
            spark_auto_username = os.environ.get('SPARK_AUTO_USERNAME')
            spark_auto_password = os.environ.get('SPARK_AUTO_PASSWORD')
            subprocess.call([
                "powershell",
                "-ExecutionPolicy", "Bypass",
                "-File", script_file,
                "-UserName", spark_auto_username,
                "-Password", spark_auto_password,
                "-JsonFilePath", submission_file,
                "-submission_MV", "true",
                "-OutputDirectory", output_directory
            ])
            logging.info(f"Successfully submitted {submission_file} to Spark using submission_MV.ps1")
        except Exception as e:
            logging.error(f"Error submitting to Spark: {e}")
            raise

    def process_submissions(self):
        # Get the path to the temporary file from command-line arguments
        data_entries = self.read_temp_file()

        # Group entries by submission type
        submission_groups = {}
        for data_entry in data_entries:
            submission_option = data_entry['SubmissionOption']
            if submission_option not in submission_groups:
                submission_groups[submission_option] = []
            submission_groups[submission_option].append(data_entry)

        # Process each group
        for submission_type, entries in submission_groups.items():
            print(f"Creating submission for {self.tp_site_excel} - {submission_type}")
            self.create_submission_json(entries, submission_type)
            self.submit_to_spark(os.path.join(self.vpo_run_directory, f'submission_vpo_{self.tp_site_excel}_{submission_type}.json'))


class GetVpoNumFromSpark:
    def __init__(self, report_file_path, submission_file_path, vpo_run_directory, tp_site_excel, script_path):
        self.report_file_path = report_file_path
        self.submission_file_path = submission_file_path
        self.vpo_run_directory = vpo_run_directory
        self.tp_site_excel = tp_site_excel  # TPSite from Excel
        self.script_path = script_path

    # TODO: update the ps file to have multiple account selection based on site
    def run_get_data_script(self):

        # Call the PowerShell script to get VPO numbers from SPARK
        logging.info(f"Calling the GetData PowerShell script for Excel TPSite {self.tp_site_excel}...")

        # Use site prefix based on Excel TPSite
        script_file = os.path.normpath(os.path.join(self.script_path, "submission_MV.ps1"))

        try:
            spark_auto_username = os.environ.get('SPARK_AUTO_USERNAME')
            spark_auto_password = os.environ.get('SPARK_AUTO_PASSWORD')
            subprocess.call([
                "powershell",
                "-ExecutionPolicy", "Bypass",
                "-File", script_file,
                "-SubmissionFilePath", self.submission_file_path,
                "-UserName", spark_auto_username,
                "-Password", spark_auto_password,
                "-GetData", 'true',
                "-ReportFilePath", self.report_file_path,
                "-OutputDirectory", self.vpo_run_directory
            ])
            logging.info(f"Successfully retrieved VPO data for Excel TPSite {self.tp_site_excel}")
        except Exception as e:
            logging.error(f"Error retrieving VPO data for Excel TPSite {self.tp_site_excel}: {e}")
            raise


class ProcessOutput:
    def __init__(self, vpo_run_directory):
        self.vpo_run_directory = vpo_run_directory
        self.output_file_path = os.path.join(self.vpo_run_directory, 'output.txt')

    def clean_line(self, line):
        return ''.join(line.replace('\x00', '').split())

    def separate_output(self):
        """Separate output.txt into correlation_table.txt and mv_table.txt"""
        correlation_file_path = os.path.join(self.vpo_run_directory, 'correlation_table.txt')
        mv_file_path = os.path.join(self.vpo_run_directory, 'mv_table.txt')

        try:
            with open(self.output_file_path, 'r', encoding='utf8', errors='ignore') as file:
                data = file.read().splitlines()

            # Clean the data
            data = [self.clean_line(line) for line in data if line.strip()]
            logging.info(f"Read {len(data)} lines from output.txt")

        except Exception as e:
            logging.error(f"Error reading output file: {e}")
            return

        with open(correlation_file_path, 'w', encoding='utf8') as correlation_file, \
             open(mv_file_path, 'w', encoding='utf8') as mv_file:

            # Write headers to both files
            for line in data:
                if line.startswith('TestProgram:') or line.startswith('ContactEmail:'):
                    correlation_file.write(line + '\n')
                    mv_file.write(line + '\n')

            # Separate data by submission type
            for line in data:
                if line.startswith('DisplayName:Correlation') or line.startswith('DisplayName:PreFuse') or line.startswith('DisplayName:PostFuse'):
                    correlation_file.write(line + '\n')
                elif line.startswith('DisplayName:MV') or line.startswith('DisplayName:MVPFused'):
                    mv_file.write(line + '\n')

        logging.info(f"Separated output into {correlation_file_path} and {mv_file_path}")

    def process_output(self):
        """Main processing method"""
        logging.info("Processing output.txt and separating into correlation and MV tables.")
        self.separate_output()
        logging.info("Output processing completed.")


class EmailVpoNumber:
    def __init__(self, vpo_run_directories):
        # Accept either a single directory (string) or list of directories
        if isinstance(vpo_run_directories, str):
            self.vpo_run_directories = [vpo_run_directories]  # Convert single to list
        else:
            self.vpo_run_directories = vpo_run_directories  # Already a list

    def read_file(self, file_path):
        try:
            with open(file_path, 'r', encoding='utf8', errors='ignore') as file:
                data = file.readlines()

            data = [line.replace('\x00', '') for line in data]
            data = [line.strip() for line in data if line.strip()]

            return data
        except Exception as e:
            logging.error(f"Error reading file {file_path}: {e}")
            return []

    def extract_emails(self, data):
        emails = set()
        for line in data:
            if line.startswith('ContactEmail:'):
                contact_emails = line.split('ContactEmail:')[1].strip()
                emails.update(contact_emails.split(','))
        return list(emails)

    def extract_test_program(self, data):
        for line in data:
            if line.startswith('TestProgram:'):
                return line.split('TestProgram:')[1].strip()
        return None

    def format_display_name(self, value):
        value = value.replace('[', ' [').replace(']', '] ')
        if len(value) > 6:
            value = value[:-6] + ' ' + value[-6:]
        if len(value) > 2:
            value = value[:-2] + ' ' + value[-2:]
        return value

    def create_table(self, data, header, table_name):
        if all(line.startswith('TestProgram:') or line.startswith('ContactEmail:') for line in data):
            return f"<p>There are no {table_name} submissions.</p>"

        table = "<table border='1' cellspacing='0' style='border-collapse: collapse;font-family: Calibri;'>\n"
        table += "  <tr>\n"
        for column in header:
            table += "      <th style='padding-right: 15px; text-align: left; background-color: #000000; color: #FFFFFF;'>{0}</th>\n".format(column.strip())
        table += "  </tr>\n"

        for row in data:
            if not row.startswith('TestProgram:') and not row.startswith('ContactEmail:'):
                columns = row.split('--')
                table += "  <tr style='margin: 0px; background-color: #DDEBF7;'>\n" if data.index(row) % 2 == 0 else "  <tr style='margin: 0px; background-color: #FFFFFF;'>\n"
                for column in columns:
                    parts = column.split(':')
                    if len(parts) == 2:
                        key, value = parts
                        if 'DisplayName' in key:
                            value = self.format_display_name(value)
                        table += "      <td style='padding-right: 15px;'>{0}</td>\n".format(value.strip())
                    else:
                        logging.warning(f"Unexpected format in line: {column}")
                table += "  </tr>\n"

        table += "</table>\n"
        return table

    def collect_all_data(self):
        """Collect data from all site directories and group by submission type"""
        all_correlation_data = []
        all_mv_data = []
        all_emails = set()
        test_program = None

        for vpo_run_directory in self.vpo_run_directories:
            # Check for correlation_table.txt and mv_table.txt in each directory
            correlation_file = os.path.join(vpo_run_directory, 'correlation_table.txt')
            mv_file = os.path.join(vpo_run_directory, 'mv_table.txt')

            if os.path.exists(correlation_file):
                correlation_data = self.read_file(correlation_file)
                all_correlation_data.extend(correlation_data)
                
                # Extract emails and test program
                emails = self.extract_emails(correlation_data)
                all_emails.update(emails)
                if not test_program:
                    test_program = self.extract_test_program(correlation_data)

            if os.path.exists(mv_file):
                mv_data = self.read_file(mv_file)
                all_mv_data.extend(mv_data)
                
                # Extract emails and test program
                emails = self.extract_emails(mv_data)
                all_emails.update(emails)
                if not test_program:
                    test_program = self.extract_test_program(mv_data)

        return all_correlation_data, all_mv_data, list(all_emails), test_program

    def draft_email(self):
        """Draft email - works for both single and multi-site"""
        all_correlation_data, all_mv_data, email_addresses, test_program = self.collect_all_data()

        if not all_correlation_data and not all_mv_data:
            logging.error("No data found in any directory. Email will not be drafted.")
            return

        print(f'all_correlation_data: {all_correlation_data}')
        print(f'all_mv_data: {all_mv_data}')
        print(f'Email_addresses: {email_addresses}')
        print(f'test_program: {test_program}')

        # Filter out header lines to check for actual submission data
        correlation_submissions = [line for line in all_correlation_data 
                                 if not line.startswith('TestProgram:') and not line.startswith('ContactEmail:')]
        mv_submissions = [line for line in all_mv_data 
                         if not line.startswith('TestProgram:') and not line.startswith('ContactEmail:')]

        # Check if we have any actual submissions
        has_correlation = len(correlation_submissions) > 0
        has_mv = len(mv_submissions) > 0

        # If no submissions at all, send error email
        if not has_correlation and not has_mv:
            self.send_error_email(email_addresses, test_program)
            return

        # Create tables with Ref_VPO positioned next to New_VPO
        correlation_header = ["DisplayName", "New_VPO", "Ref_VPO", "Lot", "Qty", "Operation", "EngID", "Tags"]
        mv_header = ["DisplayName", "New_VPO", "Ref_VPO", "Lot", "Qty", "Operation", "EngID", "Tags"]

        # Determine email subject and site info based on number of sites
        site_names = []
        for directory in self.vpo_run_directories:
            # Extract site name from directory path
            path_parts = directory.split(os.sep)
            if len(path_parts) >= 3:
                site_names.append(path_parts[-3])  # Assuming structure: .../SSID/TPSite/vpo_run

        unique_sites = list(set(site_names))

        if len(unique_sites) > 1:
            subject = f"Multi-Site VPO Submission Report - Sites: {', '.join(unique_sites)}"
            site_info = f"(Sites: {', '.join(unique_sites)})"
        else:
            subject = "VPO Submission Report"
            site_info = f"(Site: {unique_sites[0]})" if unique_sites else ""

        # Build email body dynamically based on what submissions exist
        email_body = f"""
            <html style="font-family: Calibri; font-size: 0.9em;">
            <body>
            <p>Here is the VPO submission report{' from multiple sites' if len(unique_sites) > 1 else ''}. Please review the details below:</p>
            """

        # Add correlation section only if there are correlation submissions
        if has_correlation:
            correlation_table = self.create_table(all_correlation_data, correlation_header, "Correlation")
            email_body += f"""
            <p><strong>Correlation Submission {site_info}</strong></p>
            <p><em>Includes: Correlation, PreFuse, PostFuse submissions</em></p>
            <p><strong>Test Program:</strong> {test_program}</p>
            <p><strong>Email Recipients:</strong> {'; '.join(email_addresses)}</p>
            {correlation_table}
            """

        # Add MV section only if there are MV submissions
        if has_mv:
            mv_table = self.create_table(all_mv_data, mv_header, "MV")
            email_body += f"""
            <p><strong>MV Submission {site_info}</strong></p>
            <p><em>Includes: MV, MVPFused submissions</em></p>
            <p><strong>Test Program:</strong> {test_program}</p>
            <p><strong>Email Recipients:</strong> {'; '.join(email_addresses)}</p>
            {mv_table}
            """

        email_body += """
            </body></html>"""

        print(f'email_body: {email_body}')

        # Send email via SMTP
        try:
            self._send_via_smtp(email_addresses, subject, email_body)

            # Handle release notes - check ';' and take the first ';' as the release path
            release_email = os.environ.get('AUTO_RELEASE_NOTE', 'None')
            print(f'AUTO_RELEASE_NOTE: {release_email}')
            if release_email != 'None':
                # Split by semicolon (will return single item if no semicolon)
                release_paths = [path.strip() for path in release_email.split(';')]

                # Use the first path (release notes file)
                release_notes_path = release_paths[0]

                print(f'Found {len(release_paths)} path(s) in AUTO_RELEASE_NOTE')
                print(f'Using release notes path: {release_notes_path}')

                vpo_release_email = f"{os.path.dirname(release_notes_path)}/release_vpo_email.html"
                print(f'vpo release email file: {vpo_release_email}')
                if os.path.exists(vpo_release_email):
                    os.remove(vpo_release_email)
                with open(vpo_release_email, 'w') as f:
                    f.write(email_body)

            logging.info(f"{'Multi-site' if len(unique_sites) > 1 else 'Single-site'} email sent successfully via SMTP.")
        except Exception as e:
            logging.error(f"Error sending email via SMTP: {e}")
            raise

    def _send_via_smtp(self, email_addresses, subject, email_body):
        """Send email via SMTP"""
        try:
            # SMTP Configuration
            smtp_server = "smtpauth.intel.com"
            smtp_port = 587
            sender_email = os.environ.get('SPARK_AUTO_USERNAME')
            sender_password = os.environ.get('SPARK_AUTO_PASSWORD')

            # Debug: Print what we're working with
            print(f"Raw email_addresses: {email_addresses}")
            print(f"Type of email_addresses: {type(email_addresses)}")

            # Clean TO email addresses
            if isinstance(email_addresses, str):
                to_emails = [email.strip() for email in email_addresses.replace(';', ',').split(',')]
            else:
                to_emails = [str(email).strip() for email in email_addresses]

            # Filter out empty/invalid emails
            to_emails = [email for email in to_emails if email and '@' in email and email.lower() != 'nan']

            # Add CC emails to the same recipient list (no distinction in SMTP)
            cc_emails = ["sam.ren.jie.yap@intel.com", "tai.pham@intel.com", "sys_mtls_class_tp1@intel.com"]
            all_recipients = to_emails + cc_emails

            if not all_recipients:
                raise ValueError("No valid email addresses found")

            print(f"All recipients: {all_recipients}")

            # Connect to SMTP server
            smtp = smtplib.SMTP(smtp_server, smtp_port)
            smtp.starttls()
            smtp.login(sender_email, sender_password)

            # Create the email message
            mail = MIMEMultipart()
            mail['From'] = sender_email
            # For display purposes, keep the original format
            mail['To'] = ';'.join(to_emails)
            mail['CC'] = ';'.join(cc_emails)  # This is just for display in email clients
            mail['Subject'] = subject

            # Attach HTML body
            html_body = MIMEText(email_body, 'html')
            mail.attach(html_body)

            # Send to all recipients (SMTP doesn't distinguish TO vs CC)
            smtp.sendmail(sender_email, all_recipients, mail.as_string())
            smtp.quit()

            logging.info(f"Email sent successfully to {len(all_recipients)} recipients")
            print(f"Email sent successfully to: {', '.join(all_recipients)}")

        except Exception as e:
            logging.error(f"Failed to send email via SMTP: {e}")
            print(f"Failed to send email via SMTP: {e}")
            raise

    def send_error_email(self, email_addresses, test_program):
        """Send error email when no submissions are found"""
        try:
            subject = "VPO Submission Error - No Submissions Found"
            
            error_body = f"""
            <html style="font-family: Calibri; font-size: 0.9em;">
            <body>
            <p><strong style="color: red;">ERROR: Something is wrong, please take a look</strong></p>
            <p>No VPO submissions were found for test program: <strong>{test_program}</strong></p>
            <p>Please check the submission process and try again.</p>
            <p><strong>Email Recipients:</strong> {'; '.join(email_addresses)}</p>
            </body></html>"""

            self._send_via_smtp(email_addresses, subject, error_body)

            logging.error("Error email sent - no submissions found")
            print("Error email sent - no submissions found")
            
        except Exception as e:
            logging.error(f"Error sending error email: {e}")
            print(f"Error sending error email: {e}")


class SparkAutoSubmission:
    def __init__(self, excel_file_path, current_path, script_path):
        self.excel_file_path = excel_file_path
        self.current_path = current_path
        self.script_path = script_path
        self.status = SiteProcessingStatus()

    def count_vpo_entries_in_output(self, output_file_path):
        """Count the number of VPO entries in the output file"""
        try:
            with open(output_file_path, 'r', encoding='utf-16') as f:
                data = f.read()
            
            # Count DisplayName entries (each represents a VPO)
            vpo_count = data.count('DisplayName:')
            logging.info(f"Found {vpo_count} VPO entries in output file")
            return vpo_count
        except Exception as e:
            logging.error(f"Error counting VPO entries: {e}")
            return 0

    def process(self):
        # Initialize SubmissionRead class and process the Excel file
        submission_read = SubmissionRead(self.excel_file_path, self.current_path)
        all_site_data = submission_read.process_excel()

        if not all_site_data:
            print("No site data to process.")
            return

        # VALIDATE EMAILS FROM LOG FILES AFTER PROCESSING
        print("=== Validating Contact Emails from Log Files ===")
        email_validator = EmailValidator()
        validation_errors = email_validator.validate_contact_emails_from_logs(all_site_data)

        if validation_errors:
            email_validator.prompt_for_email_correction(validation_errors)
            return  # This won't be reached due to sys.exit(1) in prompt_for_email_correction

        print("Email validation passed - proceeding with SPARK submission")
        print()

        all_vpo_run_directories = []

        # PHASE 1: Complete all submissions first
        print(f"\n=== PHASE 1: Submitting to SPARK for {len(all_site_data)} sites ===")
        for site_data in all_site_data:
            temp_file_path = site_data['temp_file_path']
            vpo_run_directory = site_data['vpo_run_directory']
            tp_site_excel = site_data['tp_site']  # TPSite from Excel
            
            try:
                print(f"\n=== Submitting Excel TPSite: {tp_site_excel} ===")

                # Initialize SparkGenApi with Excel TPSite only
                spark_gen_api = SparkGenApi(
                    temp_file_path, 
                    vpo_run_directory,
                    tp_site_excel,  # Excel TPSite
                    self.script_path
                )
                spark_gen_api.process_submissions()

                # Initialize GenerateVpoJobsReport class and generate the report
                report_generator = GenerateVpoJobsReport(os.path.join(vpo_run_directory, 'submission_log.txt'), vpo_run_directory)
                report_generator.generate_report_info()

                print(f"Successfully submitted site: {tp_site_excel}")

            except Exception as e:
                self.status.add_failure(tp_site_excel, str(e))
                print(f"Failed to submit site {tp_site_excel}: {e}")
                logging.error(f"Failed to submit site {tp_site_excel}: {e}")
                continue  # Continue with other sites even if one fails

        # PHASE 2: Get VPO data for all successfully submitted sites
        print(f"\n=== PHASE 2: Getting VPO data for submitted sites ===")
        for site_data in all_site_data:
            temp_file_path = site_data['temp_file_path']
            vpo_run_directory = site_data['vpo_run_directory']
            tp_site_excel = site_data['tp_site']  # TPSite from Excel
            expected_vpo_count = site_data['expected_vpo_count']  # Get expected count

            # Skip if this site failed in Phase 1
            if tp_site_excel in [f['site'] for f in self.status.failed_sites]:
                print(f"Skipping VPO data retrieval for failed site: {tp_site_excel}")
                continue

            # Skip it if the submission failed. No submission_id.txt found.
            submission_id_file = os.path.join(vpo_run_directory, 'submission_id.txt')
            if not os.path.exists(submission_id_file):
                print(f"Skipping VPO data retrieval for site with no submission ID: {tp_site_excel}")
                self.status.add_failure(tp_site_excel, "No submission_id.txt found, skipping VPO data retrieval")
                continue

            try:
                print(f"\n=== Getting VPO data for Excel TPSite: {tp_site_excel} ===")
                print(f"Expected VPO count for {tp_site_excel}: {expected_vpo_count}")

                # Initialize GetVpoNumFromSpark with Excel TPSite only
                retry_vpo_data_flag = True
                time_allow = 0
                max_retries = 15  # Maximum number of retry attempts
                
                while retry_vpo_data_flag and time_allow < max_retries:
                    if time_allow == 0:
                        print(f'Getting VPO data for {tp_site_excel} try# {time_allow + 1}. Waiting for 5 mins...')
                        time.sleep(300)
                    elif time_allow > 0 and time_allow < 10:
                        print(f'Getting VPO data for {tp_site_excel} try# {time_allow + 1}. Waiting for 3 mins...')
                        time.sleep(180)
                    else:
                        print(f'Getting VPO data for {tp_site_excel} try# {time_allow + 1}. Waiting for 2 mins...')
                        time.sleep(120)
                    
                    report_file_path = os.path.join(vpo_run_directory, 'report_info.txt')
                    submission_file_path = os.path.join(vpo_run_directory, 'submission_id.txt')
                    vpo_output_file_path = os.path.join(vpo_run_directory, 'output.txt')
                    
                    get_vpo_num_from_spark = GetVpoNumFromSpark(
                        report_file_path, 
                        submission_file_path, 
                        vpo_run_directory, 
                        tp_site_excel,  # Excel TPSite
                        self.script_path
                    )
                    get_vpo_num_from_spark.run_get_data_script()
                    
                    if os.path.exists(vpo_output_file_path):
                        with open(vpo_output_file_path, 'r', encoding='utf-16') as f:
                            data = f.read()

                        # Count actual VPO entries received
                        actual_vpo_count = self.count_vpo_entries_in_output(vpo_output_file_path)

                        print(f'VPO output data for {tp_site_excel}: Expected={expected_vpo_count}, Actual={actual_vpo_count}')

                        # Check if we have all expected VPOs
                        if 'Ref_VPO' in data and actual_vpo_count >= expected_vpo_count:
                            print(f'Found complete VPO output data for {tp_site_excel} ({actual_vpo_count}/{expected_vpo_count} VPOs).')
                            retry_vpo_data_flag = False  # Exit loop - SUCCESS
                        elif 'Ref_VPO' in data and actual_vpo_count < expected_vpo_count:
                            print(f'Partial VPO data for {tp_site_excel} ({actual_vpo_count}/{expected_vpo_count} VPOs). Continuing to wait...')
                            # Continue waiting, but increment counter
                        else:
                            print(f'No VPO data found for {tp_site_excel}. Continuing to wait...')
                            # Continue waiting, but increment counter
                    else:
                        print(f'VPO output file not generated for {tp_site_excel}: {vpo_output_file_path}')
                        # Continue waiting, but increment counter

                    # Always increment the retry counter
                    time_allow += 1

                # Check final status after loop exits
                if time_allow >= max_retries:
                    # Timeout reached - proceed with whatever data we have
                    final_count = self.count_vpo_entries_in_output(vpo_output_file_path) if os.path.exists(vpo_output_file_path) else 0
                    if final_count > 0:
                        print(f'TIMEOUT: Proceeding with partial data for {tp_site_excel} ({final_count}/{expected_vpo_count} VPOs)')
                        logging.warning(f"Timeout reached for {tp_site_excel}. Proceeding with {final_count}/{expected_vpo_count} VPOs")
                    else:
                        print(f'TIMEOUT: No VPO data received for {tp_site_excel} after {max_retries} attempts')
                        logging.error(f"No VPO data received for {tp_site_excel} after timeout")
                        raise Exception(f"No VPO data received for {tp_site_excel} after {max_retries} retry attempts")

                # Process output files (even if partial)
                if os.path.exists(vpo_output_file_path):
                    process_output = ProcessOutput(vpo_run_directory)
                    process_output.process_output()

                    # Create timestamped ZIP archive for this site
                    self.create_archive(vpo_run_directory, site_data['ssid'], tp_site_excel)

                    # Mark as successful (even if partial data)
                    self.status.add_success(tp_site_excel, vpo_run_directory)
                    print(f"Successfully processed site: {tp_site_excel}")
                    
                    all_vpo_run_directories.append(vpo_run_directory)
                else:
                    raise Exception(f"No output file generated for {tp_site_excel}")

            except Exception as e:
                self.status.add_failure(tp_site_excel, str(e))
                print(f"Failed to get VPO data for site {tp_site_excel}: {e}")
                logging.error(f"Failed to get VPO data for site {tp_site_excel}: {e}")

        # Print processing summary
        summary = self.status.get_summary()
        print(f"\n=== Processing Summary ===")
        print(f"Successful sites: {summary['successful_count']} - {summary['successful_sites']}")
        print(f"Failed sites: {summary['failed_count']} - {summary['failed_sites']}")

        # Send email using the unified EmailVpoNumber class (only for successful sites)
        if all_vpo_run_directories:
            email_vpo_number = EmailVpoNumber(all_vpo_run_directories)  # Pass list of directories
            email_vpo_number.draft_email()
            # After drafting/sending email, update Ref_VPO in Excel to the new VPO values
            try:
                vpo_mapping = self._build_ref_to_new_vpo_mapping(all_vpo_run_directories)
                if vpo_mapping:
                    self._update_excel_ref_vpo(self.excel_file_path, vpo_mapping)
                    print(f"Updated Ref_VPO to new VPO in Excel for {len(vpo_mapping)} entries")
                else:
                    print("No VPO mapping found to update in Excel (mapping empty)")
            except Exception as e:
                logging.error(f"Failed to update Excel Ref_VPO with new VPO values: {e}")
                print(f"Warning: Failed to update Excel Ref_VPO: {e}")
        else:
            print("No successful sites to send email for.")
            raise Exception("No successful sites to send email for.")

    def _build_ref_to_new_vpo_mapping(self, vpo_run_directories):
        """Build a mapping of Ref_VPO -> New_VPO by parsing the separated tables.
        It scans correlation_table.txt and mv_table.txt in each vpo_run directory.
        """
        mapping = {}
        def read_lines(path):
            try:
                with open(path, 'r', encoding='utf8', errors='ignore') as f:
                    return [line.strip() for line in f.readlines() if line.strip()]
            except Exception:
                return []

        for vdir in vpo_run_directories:
            for fname in ('correlation_table.txt', 'mv_table.txt'):
                fpath = os.path.join(vdir, fname)
                lines = read_lines(fpath)
                for line in lines:
                    if line.startswith('TestProgram:') or line.startswith('ContactEmail:'):
                        continue
                    # Each data row is like: key:value--key:value--...
                    parts = [p for p in line.split('--') if p]
                    kv = {}
                    for p in parts:
                        if ':' in p:
                            k, v = p.split(':', 1)
                            kv[k.strip()] = v.strip()
                    ref = kv.get('Ref_VPO')
                    # Some outputs use 'VPO' instead of 'New_VPO'
                    new = kv.get('New_VPO') or kv.get('VPO')
                    if ref and new and ref.upper() != 'N/A' and new.upper() != 'N/A':
                        mapping[ref] = new
        return mapping

    def _update_excel_ref_vpo(self, excel_file_path, ref_to_new_mapping):
        """Update the Ref_VPO column in the VPO_Jobs sheet based on mapping."""
        wb = load_workbook(excel_file_path)
        if 'VPO_Jobs' not in wb.sheetnames:
            raise ValueError("VPO_Jobs sheet not found in Excel file")
        ws = wb['VPO_Jobs']

        # Find the Ref_VPO column by header name in first row
        header_index = {}
        for cell in ws[1]:
            header_index[str(cell.value).strip() if cell.value else ''] = cell.column
        if 'Ref_VPO' not in header_index:
            raise ValueError("Ref_VPO column not found in VPO_Jobs sheet")
        ref_col = header_index['Ref_VPO']

        # Iterate data rows and replace values
        updates = 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            cell = row[ref_col - 1]
            old_val = cell.value
            if old_val is None:
                continue
            old_str = str(old_val).strip()
            if not old_str or old_str.upper() == 'N/A':
                continue
            new_vpo = ref_to_new_mapping.get(old_str)
            if new_vpo:
                cell.value = new_vpo
                updates += 1

        if updates:
            wb.save(excel_file_path)
        wb.close()

    def create_archive(self, vpo_run_directory, ssid, tp_site_excel):
        """Create timestamped ZIP archive and move all files"""
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        zip_directory = os.path.join(self.current_path, 'VPO_Description', 'VPO_Jobs', ssid, tp_site_excel, timestamp)
        print(f'Create archive at: {zip_directory}')
        os.makedirs(zip_directory, exist_ok=True)

        # Copy all files from vpo_run_directory to timestamped directory
        for file_name in os.listdir(vpo_run_directory):
            src_file = os.path.join(vpo_run_directory, file_name)
            if os.path.isfile(src_file):
                shutil.copy(src_file, zip_directory)
                logging.debug(f"Copied {src_file} to {zip_directory}")

        # Create ZIP archive
        archive_path = shutil.make_archive(zip_directory, 'zip', zip_directory)
        logging.info(f"Created archive: {archive_path}")

        # Clean up the temporary directory after archiving
        shutil.rmtree(zip_directory)
        logging.debug(f"Cleaned up temporary directory: {zip_directory}")


class SparkAutomationSelector:
    def __init__(self, excel_file_path, current_path, directory_path=None, script_path=None):
        self.excel_file_path = excel_file_path
        self.current_path = current_path
        self.directory_path = directory_path
        self.script_path = script_path

    def run(self):
        if self.directory_path:
            self.run_single_tp(self.script_path)
        else:
            # Check the Excel file for non-blank directory paths
            df_jobs = pd.read_excel(self.excel_file_path, sheet_name='VPO_Jobs')
            if df_jobs['DirectoryPath'].notna().any():
                self.run_multi_tp(self.script_path)
            else:
                raise ValueError("Error: No UserInput TP Path in both YAML file and Excel file. Please add to either one to proceed.")

    def run_single_tp(self, script_path):
        # Check for the required number of arguments
        if not self.directory_path or not self.excel_file_path or not self.current_path or not self.script_path:
            raise ValueError("Four arguments are required for SingleTP processing: ExcelFilePath, current_path, script_path, and DirectoryPath.")

        # Process user input to determine TP_REV and update the Excel file
        user_input_processor = ProcessUserInput(self.directory_path)
        tp_rev = user_input_processor.determine_tp_rev()
        user_input_processor.update_excel(self.excel_file_path, self.directory_path, tp_rev)

        # Process the submission using SparkAutoSubmission (handles both single and multi-site)
        spark_auto_submission = SparkAutoSubmission(self.excel_file_path, self.current_path, script_path)
        spark_auto_submission.process()

    def run_multi_tp(self, script_path):
        # Check for the required number of arguments
        if not self.excel_file_path or not self.current_path or not self.script_path:
            raise ValueError("Three arguments are required for MultiTP processing: ExcelFilePath, current_path, and script_path.")

        # Process the Excel file to update TPRevision
        process_multi_tp_excel = ProcessMultiTPExcel(self.excel_file_path)
        process_multi_tp_excel.process_excel_multi()

        # Process the submission using SparkAutoSubmission (handles both single and multi-site)
        spark_auto_submission = SparkAutoSubmission(self.excel_file_path, self.current_path, script_path)
        spark_auto_submission.process()


def main():
    configure_proxy()
    ensure_runtime_dependencies()

    # Get arguments without tp_site from YAML
    excel_file_path = sys.argv[1]
    current_path = sys.argv[2]
    script_path = sys.argv[3]
    directory_path = sys.argv[4] if len(sys.argv) > 4 else None

    print(f"=== Spark Automation Started ===")
    print(f"Excel file: {excel_file_path}")
    print(f"Current path: {current_path}")
    print(f"Script path: {script_path}")
    print(f"Directory path: {directory_path}")
    print()

    spark_automation_selector = SparkAutomationSelector(excel_file_path, current_path, directory_path, script_path)
    spark_automation_selector.run()

    # Shut down logging to release file handles
    logging.shutdown()

    # Clean up all vpo_run directories after archiving
    try:
        df_static = pd.read_excel(excel_file_path, sheet_name='StaticInputs')
        for _, row in df_static.iterrows():
            ssid = row['SSID']
            tp_site = row['TPSite']
            vpo_run_directory = os.path.normpath(os.path.join(current_path, 'Submission_History', ssid, tp_site, 'vpo_run'))
            if os.path.exists(vpo_run_directory):
                shutil.rmtree(vpo_run_directory)
                print(f"Cleaned up the vpo_run_directory: {vpo_run_directory}")
    except Exception as e:
        print(f"Warning: Error during cleanup: {e}")

    print("=== Spark Automation Completed ===")


if __name__ == "__main__":
    main()
